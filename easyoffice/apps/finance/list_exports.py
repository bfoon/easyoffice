"""
apps/finance/list_exports.py
============================

CSV / XLSX exporters for the finance *list* screens:

    * Invoices (receivables)      – IncomingPaymentRequest  → export_invoices_*
    * Payment Requests (payables) – PaymentRequest          → export_payment_requests_*
    * Payments (disbursements)    – Payment                 → export_payments_*

Each exporter takes an already-filtered queryset (so the download honours the
same search/status/method/year filters the user is looking at) and returns an
HttpResponse with the right Content-Disposition header.

Library choices match apps/finance/reports.py:
    * Excel: openpyxl
    * CSV:   stdlib csv

Design
------
The two format writers (`_csv_response`, `_xlsx_response`) are generic: give
them a list of column headers and a list of row tuples and they build the file.
Each entity has one `*_rows()` function that turns a queryset into
(headers, rows), so CSV and XLSX always stay in lock-step.
"""
from __future__ import annotations

import csv
from decimal import Decimal
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Generic writers ──────────────────────────────────────────────────────────

def _timestamp():
    return timezone.now().strftime('%Y%m%d_%H%M')


def _csv_response(filename: str, headers: list[str], rows: list[tuple]):
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(['' if v is None else v for v in row])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return resp


# Columns whose values are money and should carry a numeric Excel format.
def _xlsx_response(filename: str, title: str, headers: list[str], rows: list[tuple],
                   *, money_cols: set[int] = frozenset(), col_widths: dict | None = None):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet-name limit

    title_font  = Font(bold=True, size=14, color='0F172A')
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F172A')
    money_fmt   = '#,##0.00'

    # Title + generated-at banner
    ncols = max(len(headers), 1)
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=2, column=1,
            value=f'Generated {timezone.now().strftime("%Y-%m-%d %H:%M")}  ·  {len(rows)} row(s)')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    # Header row
    header_row = 4
    for c, head in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=head)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    for r, row in enumerate(rows, start=header_row + 1):
        for c, value in enumerate(row, start=1):
            if isinstance(value, Decimal):
                value = float(value)
            cell = ws.cell(row=r, column=c, value=value)
            if c in money_cols:
                cell.number_format = money_fmt

    # Freeze header, add autofilter
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = (
        f'{get_column_letter(1)}{header_row}:'
        f'{get_column_letter(ncols)}{header_row + len(rows)}'
    )

    # Column widths
    widths = col_widths or {}
    for c, head in enumerate(headers, start=1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = widths.get(c, max(12, min(40, len(str(head)) + 6)))

    buf = BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return resp


# ── Small helpers ────────────────────────────────────────────────────────────

def _d(value) -> Decimal:
    try:
        return Decimal(str(value or '0'))
    except Exception:
        return Decimal('0')


def _date(value):
    return value.strftime('%Y-%m-%d') if value else ''


def _display(obj, field):
    """Prefer a get_<field>_display() label if the model provides one."""
    getter = getattr(obj, f'get_{field}_display', None)
    if callable(getter):
        try:
            return getter()
        except Exception:
            pass
    return getattr(obj, field, '') or ''


# ════════════════════════════════════════════════════════════════════════════
# 1. INVOICES  (IncomingPaymentRequest — receivables / outgoing invoices)
# ════════════════════════════════════════════════════════════════════════════

def _invoice_total(ipr) -> Decimal:
    """Use the model's total if present, else amount + tax − discount."""
    total = getattr(ipr, 'total_amount', None)
    if total is not None:
        return _d(total)
    return _d(getattr(ipr, 'amount', 0)) + _d(getattr(ipr, 'tax_amount', 0)) - _d(getattr(ipr, 'discount_amount', 0))


def invoices_rows(qs):
    headers = [
        'Invoice #', 'Title', 'Customer', 'Company', 'Email',
        'Issue Date', 'Due Date', 'Status',
        'Currency', 'Amount', 'Tax', 'Discount', 'Total',
        'Paid At', 'Reminders', 'Last Reminder',
    ]
    money_cols = {10, 11, 12, 13}  # Amount, Tax, Discount, Total (1-indexed)
    rows = []
    for ipr in qs:
        rows.append((
            getattr(ipr, 'invoice_number', ''),
            getattr(ipr, 'title', ''),
            getattr(ipr, 'customer_name', ''),
            getattr(ipr, 'customer_company', '') or '',
            getattr(ipr, 'customer_email', '') or '',
            _date(getattr(ipr, 'issue_date', None)),
            _date(getattr(ipr, 'due_date', None)),
            _display(ipr, 'status'),
            getattr(ipr, 'currency', '') or '',
            _d(getattr(ipr, 'amount', 0)),
            _d(getattr(ipr, 'tax_amount', 0)),
            _d(getattr(ipr, 'discount_amount', 0)),
            _invoice_total(ipr),
            _date(getattr(ipr, 'paid_at', None)),
            getattr(ipr, 'reminder_count', 0) or 0,
            _date(getattr(ipr, 'last_reminder_sent', None)),
        ))
    return headers, rows, money_cols


def export_invoices(qs, fmt: str = 'xlsx'):
    headers, rows, money_cols = invoices_rows(qs)
    name = f'invoices_{_timestamp()}'
    if fmt == 'csv':
        return _csv_response(name, headers, rows)
    return _xlsx_response(
        name, 'Invoices', headers, rows,
        money_cols=money_cols,
        col_widths={1: 18, 2: 30, 3: 24, 4: 24, 5: 26, 13: 16},
    )


# ════════════════════════════════════════════════════════════════════════════
# 2. PAYMENT REQUESTS  (PaymentRequest — payables / outgoing payment requests)
# ════════════════════════════════════════════════════════════════════════════

def payment_requests_rows(qs):
    headers = [
        'Title', 'Recipient', 'Recipient Type', 'Email',
        'Currency', 'Amount', 'Status', 'Payment Type',
        'Due Date', 'Budget', 'Project', 'Requested By',
        'Created', 'Linked Payment Ref',
    ]
    money_cols = {6}
    rows = []
    for pr in qs:
        linked = getattr(pr, 'linked_payment', None)
        rows.append((
            getattr(pr, 'title', ''),
            getattr(pr, 'recipient_name', '') or (
                getattr(getattr(pr, 'recipient_user', None), 'get_full_name', lambda: '')()
                if getattr(pr, 'recipient_user', None) else ''
            ),
            _display(pr, 'recipient_type'),
            getattr(pr, 'recipient_email', '') or '',
            getattr(pr, 'currency', '') or '',
            _d(getattr(pr, 'amount', 0)),
            _display(pr, 'status'),
            _display(pr, 'payment_type'),
            _date(getattr(pr, 'due_date', None)),
            getattr(getattr(pr, 'budget', None), 'name', '') or '',
            getattr(getattr(pr, 'project', None), 'name', '') or '',
            getattr(getattr(pr, 'requested_by', None), 'get_full_name', lambda: '')() if getattr(pr, 'requested_by', None) else '',
            _date(getattr(pr, 'created_at', None)),
            getattr(linked, 'reference', '') if linked else '',
        ))
    return headers, rows, money_cols


def export_payment_requests(qs, fmt: str = 'xlsx'):
    headers, rows, money_cols = payment_requests_rows(qs)
    name = f'payment_requests_{_timestamp()}'
    if fmt == 'csv':
        return _csv_response(name, headers, rows)
    return _xlsx_response(
        name, 'Payment Requests', headers, rows,
        money_cols=money_cols,
        col_widths={1: 30, 2: 24, 4: 26, 10: 20, 11: 20},
    )


# ════════════════════════════════════════════════════════════════════════════
# 3. PAYMENTS  (Payment — disbursements / transaction history)
# ════════════════════════════════════════════════════════════════════════════

def payments_rows(qs):
    headers = [
        'Reference', 'Recipient', 'Description',
        'Currency', 'Amount', 'Method', 'Direction', 'Payment Type',
        'Date', 'Budget', 'Project', 'Purchase Request', 'Notes',
    ]
    money_cols = {5}
    rows = []
    for p in qs:
        rows.append((
            getattr(p, 'reference', ''),
            getattr(p, 'recipient', '') or '',
            getattr(p, 'description', '') or '',
            getattr(p, 'currency', '') or '',
            _d(getattr(p, 'amount', 0)),
            _display(p, 'method'),
            _display(p, 'direction'),
            _display(p, 'payment_type'),
            _date(getattr(p, 'payment_date', None)),
            getattr(getattr(p, 'budget', None), 'name', '') or '',
            getattr(getattr(p, 'project', None), 'name', '') or '',
            getattr(getattr(p, 'purchase_request', None), 'title', '') or '',
            (getattr(p, 'notes', '') or '')[:500],
        ))
    return headers, rows, money_cols


def export_payments(qs, fmt: str = 'xlsx'):
    headers, rows, money_cols = payments_rows(qs)
    name = f'payments_{_timestamp()}'
    if fmt == 'csv':
        return _csv_response(name, headers, rows)
    return _xlsx_response(
        name, 'Payments', headers, rows,
        money_cols=money_cols,
        col_widths={1: 18, 2: 24, 3: 34, 12: 28, 13: 40},
    )
