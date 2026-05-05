"""
apps/finance/reports.py
========================

PDF and Excel exporters for the five core reports:

  1. Income Statement (P&L)
  2. Cash Flow Statement
  3. Budget vs Actual
  4. Audit Trail / Change Log
  5. Ageing (Receivables & Payables)

Library choices match the rest of EasyOffice:
- PDFs: reportlab (already used by the invoices app)
- Excel: openpyxl

Public API
----------
Each export returns an HttpResponse with the right Content-Disposition header.
Views call these directly:

    return export_pl_pdf(period, basis)
    return export_pl_xlsx(period, basis)
    ...
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.finance import analytics
from apps.finance.models import FinanceAuditLog


# ─── Common styling ──────────────────────────────────────────────────────────

BRAND_NAVY   = colors.HexColor('#0f172a')
BRAND_GOLD   = colors.HexColor('#b08d57')
BRAND_GREY   = colors.HexColor('#475569')
LIGHT_GREY   = colors.HexColor('#f1f5f9')
BORDER_GREY  = colors.HexColor('#cbd5e1')
GREEN        = colors.HexColor('#059669')
RED          = colors.HexColor('#dc2626')


def _money(v, currency='GMD'):
    if v is None:
        return '—'
    if isinstance(v, Decimal):
        v = float(v)
    sign = '-' if v < 0 else ''
    return f'{sign}{currency} {abs(v):,.2f}'


def _styles():
    s = getSampleStyleSheet()
    s.add(ParagraphStyle(
        name='ReportTitle', fontName='Helvetica-Bold', fontSize=20,
        textColor=BRAND_NAVY, leading=24, spaceAfter=4,
    ))
    s.add(ParagraphStyle(
        name='ReportSubtitle', fontName='Helvetica', fontSize=10,
        textColor=BRAND_GREY, leading=13, spaceAfter=18,
    ))
    s.add(ParagraphStyle(
        name='SectionHeader', fontName='Helvetica-Bold', fontSize=12,
        textColor=BRAND_NAVY, leading=16, spaceBefore=14, spaceAfter=6,
    ))
    s.add(ParagraphStyle(
        name='Footer', fontName='Helvetica', fontSize=8,
        textColor=BRAND_GREY, alignment=TA_CENTER,
    ))
    return s


def _header(title, subtitle, styles):
    return [
        Paragraph(title, styles['ReportTitle']),
        Paragraph(subtitle, styles['ReportSubtitle']),
    ]


def _money_table(rows, label_w=110, value_w=50, totals_idx=()):
    """Build a clean two-column money table."""
    data = []
    for label, value in rows:
        data.append([label, value])
    t = Table(data, colWidths=[label_w * mm, value_w * mm])
    style = [
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), BRAND_NAVY),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, -1), 0.4, BORDER_GREY),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]
    for idx in totals_idx:
        style += [
            ('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'),
            ('BACKGROUND', (0, idx), (-1, idx), LIGHT_GREY),
            ('LINEBELOW', (0, idx), (-1, idx), 1.2, BRAND_NAVY),
        ]
    t.setStyle(TableStyle(style))
    return t


# ════════════════════════════════════════════════════════════════════════════
# 1. INCOME STATEMENT
# ════════════════════════════════════════════════════════════════════════════

def export_pl_pdf(period, basis='cash'):
    data = analytics.income_statement(period, basis)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=18*mm, bottomMargin=18*mm,
                            title=f'Income Statement {period.start} to {period.end}')
    styles = _styles()
    story = _header(
        'Income Statement',
        f'{period.label or ""}  ·  {period.start.isoformat()} → {period.end.isoformat()}'
        f'  ·  {basis.title()} basis',
        styles,
    )

    # Revenue section
    story.append(Paragraph('Revenue', styles['SectionHeader']))
    rev_rows = [(line['label'], _money(line['amount'])) for line in data['revenue_lines']]
    rev_rows.append(('Total Revenue', _money(data['revenue'])))
    story.append(_money_table(rev_rows, totals_idx=[len(rev_rows) - 1]))

    # Expenses
    story.append(Paragraph('Operating Expenses', styles['SectionHeader']))
    exp_rows = [(line['label'], _money(line['amount'])) for line in data['expense_lines']]
    if not exp_rows:
        exp_rows = [('No expenses recorded', _money(0))]
    exp_rows.append(('Total Expenses', _money(data['total_expenses'])))
    story.append(_money_table(exp_rows, totals_idx=[len(exp_rows) - 1]))

    # Net
    story.append(Paragraph('Net Result', styles['SectionHeader']))
    net_rows = [
        ('Net Income / (Loss)', _money(data['net_income'])),
        ('Margin', f'{float(data["margin_pct"]):.1f}%'),
    ]
    story.append(_money_table(net_rows, totals_idx=[0]))

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f'Generated {timezone.now().strftime("%d %b %Y %H:%M")} · '
        f'EasyOffice Financial Control Center',
        styles['Footer'],
    ))

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    fname = f'income_statement_{period.start}_{period.end}.pdf'
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_pl_xlsx(period, basis='cash'):
    data = analytics.income_statement(period, basis)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Income Statement'

    bold = Font(bold=True, color='0F172A')
    title_font = Font(bold=True, size=16, color='0F172A')
    header_fill = PatternFill('solid', fgColor='F1F5F9')
    money_fmt = '#,##0.00'

    ws['A1'] = 'Income Statement'
    ws['A1'].font = title_font
    ws.merge_cells('A1:C1')
    ws['A2'] = (
        f'{period.start} → {period.end}  ·  {basis.title()} basis  ·  '
        f'Generated {timezone.now().strftime("%Y-%m-%d %H:%M")}'
    )
    ws.merge_cells('A2:C2')

    row = 4
    ws.cell(row=row, column=1, value='REVENUE').font = bold
    ws.cell(row=row, column=1).fill = header_fill
    row += 1
    for line in data['revenue_lines']:
        ws.cell(row=row, column=1, value=line['label'])
        c = ws.cell(row=row, column=2, value=float(line['amount']))
        c.number_format = money_fmt
        row += 1
    ws.cell(row=row, column=1, value='Total Revenue').font = bold
    c = ws.cell(row=row, column=2, value=float(data['revenue']))
    c.font = bold
    c.number_format = money_fmt
    row += 2

    ws.cell(row=row, column=1, value='OPERATING EXPENSES').font = bold
    ws.cell(row=row, column=1).fill = header_fill
    row += 1
    for line in data['expense_lines']:
        ws.cell(row=row, column=1, value=line['label'])
        c = ws.cell(row=row, column=2, value=float(line['amount']))
        c.number_format = money_fmt
        row += 1
    ws.cell(row=row, column=1, value='Total Expenses').font = bold
    c = ws.cell(row=row, column=2, value=float(data['total_expenses']))
    c.font = bold
    c.number_format = money_fmt
    row += 2

    ws.cell(row=row, column=1, value='NET INCOME').font = bold
    c = ws.cell(row=row, column=2, value=float(data['net_income']))
    c.font = bold
    c.number_format = money_fmt
    row += 1
    ws.cell(row=row, column=1, value='Margin %')
    ws.cell(row=row, column=2, value=float(data['margin_pct']) / 100).number_format = '0.0%'

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 22

    return _xlsx_response(wb, f'income_statement_{period.start}_{period.end}.xlsx')


# ════════════════════════════════════════════════════════════════════════════
# 2. CASH FLOW
# ════════════════════════════════════════════════════════════════════════════

def export_cashflow_pdf(period):
    data = analytics.cash_flow_statement(period)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=18*mm, bottomMargin=18*mm,
                            title=f'Cash Flow {period.start} to {period.end}')
    styles = _styles()
    story = _header(
        'Statement of Cash Flows',
        f'{period.label or ""}  ·  {period.start.isoformat()} → {period.end.isoformat()}',
        styles,
    )

    story.append(Paragraph('Opening Cash Position', styles['SectionHeader']))
    story.append(_money_table([('Cash at start of period', _money(data['opening_cash']))]))

    story.append(Paragraph('Cash Inflows', styles['SectionHeader']))
    in_rows = [(line['label'], _money(line['amount'])) for line in data['inflows']]
    in_rows.append(('Total Inflows', _money(data['cash_in'])))
    story.append(_money_table(in_rows, totals_idx=[len(in_rows) - 1]))

    story.append(Paragraph('Cash Outflows', styles['SectionHeader']))
    out_rows = [(line['label'], _money(line['amount'])) for line in data['outflows']]
    if not out_rows:
        out_rows = [('No outflows recorded', _money(0))]
    out_rows.append(('Total Outflows', _money(data['total_outflows'])))
    story.append(_money_table(out_rows, totals_idx=[len(out_rows) - 1]))

    story.append(Paragraph('Net Cash Movement', styles['SectionHeader']))
    story.append(_money_table([
        ('Net change in cash', _money(data['net_cash_change'])),
        ('Closing cash position', _money(data['closing_cash'])),
    ], totals_idx=[1]))

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f'Generated {timezone.now().strftime("%d %b %Y %H:%M")} · '
        f'EasyOffice Financial Control Center',
        styles['Footer'],
    ))

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    fname = f'cashflow_{period.start}_{period.end}.pdf'
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_cashflow_xlsx(period):
    data = analytics.cash_flow_statement(period)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cash Flow'
    bold = Font(bold=True)
    title_font = Font(bold=True, size=16)
    header_fill = PatternFill('solid', fgColor='F1F5F9')
    money_fmt = '#,##0.00'

    ws['A1'] = 'Statement of Cash Flows'
    ws['A1'].font = title_font
    ws['A2'] = f'{period.start} → {period.end}'

    row = 4
    sections = [
        ('OPENING CASH', [('Cash at start of period', data['opening_cash'])]),
        ('INFLOWS', [(l['label'], l['amount']) for l in data['inflows']] +
                    [('Total Inflows', data['cash_in'])]),
        ('OUTFLOWS', [(l['label'], l['amount']) for l in data['outflows']] +
                     [('Total Outflows', data['total_outflows'])]),
        ('NET MOVEMENT', [
            ('Net change in cash', data['net_cash_change']),
            ('Closing cash position', data['closing_cash']),
        ]),
    ]
    for section_title, lines in sections:
        ws.cell(row=row, column=1, value=section_title).font = bold
        ws.cell(row=row, column=1).fill = header_fill
        row += 1
        for label, amt in lines:
            ws.cell(row=row, column=1, value=label)
            c = ws.cell(row=row, column=2, value=float(amt or 0))
            c.number_format = money_fmt
            if label.startswith('Total') or label.startswith('Closing') or label.startswith('Net'):
                ws.cell(row=row, column=1).font = bold
                c.font = bold
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 22
    return _xlsx_response(wb, f'cashflow_{period.start}_{period.end}.xlsx')


# ════════════════════════════════════════════════════════════════════════════
# 3. BUDGET VS ACTUAL
# ════════════════════════════════════════════════════════════════════════════

def export_budget_vs_actual_pdf(fiscal_year=None):
    data = analytics.budget_vs_actual(fiscal_year)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm,
                            title=f'Budget vs Actual FY{data["fiscal_year"]}')
    styles = _styles()
    story = _header(
        f'Budget vs Actual — FY{data["fiscal_year"]}',
        f'Snapshot at {timezone.now().strftime("%d %b %Y %H:%M")}',
        styles,
    )

    head = ['Budget', 'Department', 'Status', 'Total', 'Spent', 'Balance', 'Util %']
    rows = [head]
    for r in data['rows']:
        rows.append([
            r['name'],
            r['department'],
            r['status'].title(),
            _money(r['total']),
            _money(r['spent']),
            _money(r['balance']),
            f'{float(r["utilization"]):.1f}%',
        ])
    rows.append([
        'TOTAL', '', '',
        _money(data['grand_total']),
        _money(data['grand_spent']),
        _money(data['grand_balance']),
        f'{float(data["grand_utilization"]):.1f}%',
    ])

    t = Table(rows, repeatRows=1, colWidths=[
        55*mm, 35*mm, 22*mm, 35*mm, 35*mm, 35*mm, 22*mm,
    ])
    style = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, -1), 0.3, BORDER_GREY),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), LIGHT_GREY),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]
    # Highlight over-budget rows
    for i, r in enumerate(data['rows'], start=1):
        if r['over_budget']:
            style.append(('TEXTCOLOR', (0, i), (-1, i), RED))
    t.setStyle(TableStyle(style))
    story.append(t)

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    fname = f'budget_vs_actual_FY{data["fiscal_year"]}.pdf'
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_budget_vs_actual_xlsx(fiscal_year=None):
    data = analytics.budget_vs_actual(fiscal_year)
    wb = Workbook()
    ws = wb.active
    ws.title = f'Budget FY{data["fiscal_year"]}'

    bold = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F172A')
    money_fmt = '#,##0.00'
    pct_fmt = '0.0%'

    headers = ['Budget', 'Department', 'Unit', 'Status', 'Total', 'Spent', 'Balance', 'Utilization']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal='left' if col <= 4 else 'right')

    for i, r in enumerate(data['rows'], start=2):
        ws.cell(row=i, column=1, value=r['name'])
        ws.cell(row=i, column=2, value=r['department'])
        ws.cell(row=i, column=3, value=r['unit'])
        ws.cell(row=i, column=4, value=r['status'].title())
        ws.cell(row=i, column=5, value=float(r['total'])).number_format = money_fmt
        ws.cell(row=i, column=6, value=float(r['spent'])).number_format = money_fmt
        ws.cell(row=i, column=7, value=float(r['balance'])).number_format = money_fmt
        ws.cell(row=i, column=8, value=float(r['utilization']) / 100).number_format = pct_fmt
        if r['over_budget']:
            for col in range(1, 9):
                ws.cell(row=i, column=col).font = Font(color='DC2626')

    total_row = len(data['rows']) + 2
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=float(data['grand_total'])).number_format = money_fmt
    ws.cell(row=total_row, column=6, value=float(data['grand_spent'])).number_format = money_fmt
    ws.cell(row=total_row, column=7, value=float(data['grand_balance'])).number_format = money_fmt
    ws.cell(row=total_row, column=8, value=float(data['grand_utilization']) / 100).number_format = pct_fmt
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).fill = PatternFill('solid', fgColor='F1F5F9')

    widths = [38, 22, 18, 14, 18, 18, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return _xlsx_response(wb, f'budget_vs_actual_FY{data["fiscal_year"]}.xlsx')


# ════════════════════════════════════════════════════════════════════════════
# 4. AUDIT TRAIL
# ════════════════════════════════════════════════════════════════════════════

def export_audit_pdf(period, qs=None):
    qs = qs or FinanceAuditLog.objects.filter(
        timestamp__date__gte=period.start,
        timestamp__date__lte=period.end,
    ).select_related('actor').order_by('-timestamp')

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title='Audit Trail')
    styles = _styles()
    story = _header(
        'Financial Audit Trail',
        f'{period.start} → {period.end}  ·  {qs.count()} entries',
        styles,
    )

    head = ['Timestamp', 'Actor', 'Action', 'Model', 'Object', 'Amount', 'IP']
    rows = [head]
    for e in qs[:1000]:  # safety cap
        rows.append([
            e.timestamp.strftime('%Y-%m-%d %H:%M'),
            e.actor_name or (e.actor and e.actor.username) or 'System',
            e.get_action_display(),
            e.model_name,
            (e.object_repr or '')[:40],
            _money(e.amount, e.currency or 'GMD') if e.amount else '—',
            e.ip_address or '—',
        ])
    t = Table(rows, repeatRows=1, colWidths=[
        30*mm, 35*mm, 20*mm, 28*mm, 60*mm, 28*mm, 25*mm,
    ])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, -1), 0.2, BORDER_GREY),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
    ]))
    story.append(t)

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    fname = f'audit_trail_{period.start}_{period.end}.pdf'
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_audit_xlsx(period, qs=None):
    qs = qs or FinanceAuditLog.objects.filter(
        timestamp__date__gte=period.start,
        timestamp__date__lte=period.end,
    ).select_related('actor').order_by('-timestamp')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Audit Trail'
    bold = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F172A')
    money_fmt = '#,##0.00'

    headers = ['Timestamp', 'Actor', 'Username', 'Action', 'Model', 'Object',
               'Amount', 'Currency', 'IP', 'Changes']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = header_fill

    for i, e in enumerate(qs, start=2):
        ws.cell(row=i, column=1, value=e.timestamp.replace(tzinfo=None))
        ws.cell(row=i, column=1).number_format = 'yyyy-mm-dd hh:mm'
        ws.cell(row=i, column=2, value=e.actor_name or '')
        ws.cell(row=i, column=3, value=e.actor.username if e.actor else 'system')
        ws.cell(row=i, column=4, value=e.get_action_display())
        ws.cell(row=i, column=5, value=e.model_name)
        ws.cell(row=i, column=6, value=(e.object_repr or '')[:200])
        if e.amount is not None:
            ws.cell(row=i, column=7, value=float(e.amount)).number_format = money_fmt
        ws.cell(row=i, column=8, value=e.currency or '')
        ws.cell(row=i, column=9, value=e.ip_address or '')
        ws.cell(row=i, column=10, value=str(e.changes)[:1000] if e.changes else '')

    widths = [18, 22, 16, 14, 18, 32, 16, 10, 16, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return _xlsx_response(wb, f'audit_trail_{period.start}_{period.end}.xlsx')


# ════════════════════════════════════════════════════════════════════════════
# 5. AGEING (AR + AP)
# ════════════════════════════════════════════════════════════════════════════

def export_ageing_pdf():
    ar = analytics.receivables_ageing()
    ap = analytics.payables_ageing()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=18*mm, rightMargin=18*mm,
                            topMargin=15*mm, bottomMargin=15*mm,
                            title='Ageing Report')
    styles = _styles()
    story = _header(
        'Ageing Report',
        f'As of {ar["as_of"].strftime("%d %b %Y")}',
        styles,
    )

    def _ageing_section(title, data, col_label):
        story.append(Paragraph(title, styles['SectionHeader']))
        head = ['Bucket', 'Count', col_label]
        rows = [head]
        for b in data['buckets']:
            rows.append([b['label'], str(b['count']), _money(b['amount'])])
        rows.append(['TOTAL', '', _money(data['total'])])
        t = Table(rows, repeatRows=1, colWidths=[60*mm, 30*mm, 50*mm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), BRAND_NAVY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('LINEBELOW', (0, 0), (-1, -1), 0.3, BORDER_GREY),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), LIGHT_GREY),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(t)

    _ageing_section('Receivables (Money owed to us)', ar, 'Outstanding')
    story.append(Spacer(1, 12))
    _ageing_section('Payables (Money we owe)', ap, 'Outstanding')

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    fname = f'ageing_{ar["as_of"]}.pdf'
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def export_ageing_xlsx():
    ar = analytics.receivables_ageing()
    ap = analytics.payables_ageing()
    wb = Workbook()

    for sheet_title, data in [('Receivables', ar), ('Payables', ap)]:
        ws = wb.create_sheet(sheet_title) if sheet_title != 'Receivables' else wb.active
        if sheet_title == 'Receivables':
            ws.title = 'Receivables'
        bold = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='0F172A')
        money_fmt = '#,##0.00'

        ws['A1'] = sheet_title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'As of {data["as_of"]}'

        for col, h in enumerate(['Bucket', 'Count', 'Amount'], start=1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = bold
            c.fill = header_fill

        for i, b in enumerate(data['buckets'], start=5):
            ws.cell(row=i, column=1, value=b['label'])
            ws.cell(row=i, column=2, value=b['count'])
            ws.cell(row=i, column=3, value=float(b['amount'])).number_format = money_fmt

        total_row = 5 + len(data['buckets'])
        ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=float(data['total'])).number_format = money_fmt
        ws.cell(row=total_row, column=3).font = Font(bold=True)

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 22

    return _xlsx_response(wb, f'ageing_{ar["as_of"]}.xlsx')


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _xlsx_response(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
