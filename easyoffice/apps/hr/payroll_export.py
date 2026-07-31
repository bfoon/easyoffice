"""
apps/hr/payroll_export.py
=========================

Download the payroll register for a period as Excel, CSV or PDF.

    GET /hr/payroll/export/?year=2026&month=7&format=xlsx
    GET /hr/payroll/export/?year=2026&month=7&format=csv
    GET /hr/payroll/export/?year=2026&month=7&format=pdf
    GET /hr/payroll/export/?year=2026&month=7&format=xlsx&status=approved

Deliberately kept OUT of apps/hr/views.py so that file doesn't grow another
300 lines — this module only imports FROM views.py (the permission helpers),
never the other way round, so there's no circular import.

Same permission gate as the payroll page itself: can_run_payroll().
Excel needs openpyxl and PDF needs reportlab — both are already used
elsewhere in the project (apps/finance/reports.py, apps/hr/offer_letter.py).
CSV has no dependencies at all, so it always works as a fallback.
"""
from __future__ import annotations

import calendar
import csv
from decimal import Decimal
from io import BytesIO, StringIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.views.generic import View

from apps.hr.models import PayrollRecord
from apps.hr.views import can_run_payroll


# ── Column definition ────────────────────────────────────────────────────────
# (header, accessor, kind)   kind: 'text' | 'money' | 'days'
# Add a column here and it appears in Excel AND CSV automatically. The PDF
# uses a narrower subset (see PDF_COLUMNS) so it stays readable on A4.

def _dec(value) -> Decimal:
    if value in (None, ''):
        return Decimal('0.00')
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _staff_name(rec):
    staff = rec.staff
    return (staff.get_full_name() or staff.username) if staff else '—'


def _employee_code(rec):
    return getattr(rec.employment, 'employee_code', '') or '—' if rec.employment else '—'


def _department(rec):
    return getattr(rec.employment, 'department_display', '') or '—' if rec.employment else '—'


def _position(rec):
    return getattr(rec.employment, 'job_title_display', '') or '—' if rec.employment else '—'


def _allowances_paid(rec):
    return _dec((rec.allowances or {}).get('paid_total', 0))


def _allowances_monthly(rec):
    return _dec((rec.allowances or {}).get('monthly_total', 0))


def _payslip_sent(rec):
    sent = getattr(rec, 'payslip_sent_at', None)
    return timezone.localtime(sent).strftime('%Y-%m-%d %H:%M') if sent else '—'


def _payment_date(rec):
    d = getattr(rec, 'payment_date', None)
    return d.strftime('%Y-%m-%d') if d else '—'


COLUMNS = [
    ('#',                 lambda rec, i: i,                                  'text'),
    ('Employee Code',     lambda rec, i: _employee_code(rec),                'text'),
    ('Staff',             lambda rec, i: _staff_name(rec),                   'text'),
    ('Email',             lambda rec, i: (rec.staff.email if rec.staff else '') or '—', 'text'),
    ('Department',        lambda rec, i: _department(rec),                   'text'),
    ('Position',          lambda rec, i: _position(rec),                     'text'),
    ('Period',            lambda rec, i: f'{rec.period_month:02d}/{rec.period_year}', 'text'),
    ('Work Days',         lambda rec, i: _dec(rec.total_work_days),          'days'),
    ('Payable Days',      lambda rec, i: _dec(rec.payable_days),             'days'),
    ('Absent',            lambda rec, i: _dec(rec.absent_days),              'days'),
    ('Half Days',         lambda rec, i: _dec(rec.half_days),                'days'),
    ('Unpaid Leave',      lambda rec, i: _dec(rec.unpaid_leave_days),        'days'),
    ('Unrecorded',        lambda rec, i: _dec(rec.unrecorded_days),          'days'),
    ('Basic Salary',      lambda rec, i: _dec(rec.basic_salary),             'money'),
    ('Allowances (paid)', lambda rec, i: _allowances_paid(rec),              'money'),
    ('Leave Payout',      lambda rec, i: _dec(rec.leave_payout),             'money'),
    ('Gross Salary',      lambda rec, i: _dec(rec.gross_salary),             'money'),
    ('Tax Deducted',      lambda rec, i: _dec(rec.tax_deducted),             'money'),
    ('Net Salary',        lambda rec, i: _dec(rec.net_salary),               'money'),
    ('Status',            lambda rec, i: rec.get_status_display(),           'text'),
    ('Payment Date',      lambda rec, i: _payment_date(rec),                 'text'),
    ('Payment Reference', lambda rec, i: getattr(rec, 'payment_reference', '') or '—', 'text'),
    ('Payslip Sent',      lambda rec, i: _payslip_sent(rec),                 'text'),
]

# Indices of COLUMNS to show in the PDF (A4 landscape can't carry 23 columns).
PDF_COLUMNS = [0, 1, 2, 4, 8, 13, 14, 16, 18, 19]

MONEY_COLUMNS = [idx for idx, (_, _, kind) in enumerate(COLUMNS) if kind == 'money']


# ── Data ─────────────────────────────────────────────────────────────────────

def payroll_queryset(year: int, month: int, status: str = ''):
    """Exactly the register shown on the payroll page, optionally status-filtered."""
    qs = (
        PayrollRecord.objects
        .filter(period_year=year, period_month=month, employment__isnull=False)
        .select_related('staff', 'staff__staffprofile', 'employment',
                        'employment__department_ref', 'employment__position')
        .order_by('staff__last_name', 'staff__first_name')
    )
    if status and status.lower() != 'all':
        qs = qs.filter(status=status.lower())
    return qs


def build_rows(records):
    """[[cell, cell, …], …] — one list per payroll record, in COLUMNS order."""
    rows = []
    for i, rec in enumerate(records, start=1):
        rows.append([accessor(rec, i) for _, accessor, _ in COLUMNS])
    return rows


def build_totals(records):
    """Column-aligned totals row (money columns summed, everything else blank)."""
    totals = [''] * len(COLUMNS)
    totals[0] = 'TOTAL'
    for idx in MONEY_COLUMNS:
        accessor = COLUMNS[idx][1]
        totals[idx] = sum((accessor(rec, 0) for rec in records), Decimal('0.00'))
    return totals


def period_label(year: int, month: int) -> str:
    return f'{calendar.month_name[month]} {year}'


def export_filename(year: int, month: int, ext: str) -> str:
    return f'payroll_register_{year}_{month:02d}.{ext}'


# ── Excel ────────────────────────────────────────────────────────────────────

def export_xlsx(records, year, month) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    records = list(records)
    wb = Workbook()
    ws = wb.active
    ws.title = f'Payroll {month:02d}-{year}'

    title_font = Font(bold=True, size=16, color='0F172A')
    bold = Font(bold=True, color='0F172A')
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F3D73')
    total_fill = PatternFill('solid', fgColor='F1F5F9')
    money_fmt = '#,##0.00'
    days_fmt = '0.##'

    last_col = get_column_letter(len(COLUMNS))

    ws['A1'] = f'Payroll Register — {period_label(year, month)}'
    ws['A1'].font = title_font
    ws.merge_cells(f'A1:{last_col}1')
    ws['A2'] = (
        f'{len(records)} record(s)  ·  Generated '
        f'{timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")}'
    )
    ws.merge_cells(f'A2:{last_col}2')

    header_row = 4
    for col, (header, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row_no = header_row
    for row in build_rows(records):
        row_no += 1
        for col, (value, (_, _, kind)) in enumerate(zip(row, COLUMNS), start=1):
            if kind in ('money', 'days') and isinstance(value, Decimal):
                cell = ws.cell(row=row_no, column=col, value=float(value))
                cell.number_format = money_fmt if kind == 'money' else days_fmt
            else:
                ws.cell(row=row_no, column=col, value=value)

    if records:
        row_no += 1
        for col, value in enumerate(build_totals(records), start=1):
            if isinstance(value, Decimal):
                cell = ws.cell(row=row_no, column=col, value=float(value))
                cell.number_format = money_fmt
            else:
                cell = ws.cell(row=row_no, column=col, value=value)
            cell.font = bold
            cell.fill = total_fill

    # Column widths from the header + a sane cap
    for col, (header, _, kind) in enumerate(COLUMNS, start=1):
        width = 12 if kind in ('money', 'days') else max(11, min(len(header) + 6, 34))
        if header in ('Staff', 'Email', 'Department', 'Position'):
            width = 26
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f'A{header_row}:{last_col}{header_row + len(records)}'

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{export_filename(year, month, "xlsx")}"'
    )
    return response


# ── CSV ──────────────────────────────────────────────────────────────────────

def export_csv(records, year, month) -> HttpResponse:
    records = list(records)
    buffer = StringIO()
    writer = csv.writer(buffer)

    writer.writerow([f'Payroll Register — {period_label(year, month)}'])
    writer.writerow([
        f'{len(records)} record(s)',
        f'Generated {timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")}',
    ])
    writer.writerow([])
    writer.writerow([header for header, _, _ in COLUMNS])

    for row in build_rows(records):
        writer.writerow([
            f'{value:.2f}' if isinstance(value, Decimal) else value
            for value in row
        ])

    if records:
        writer.writerow([
            f'{value:.2f}' if isinstance(value, Decimal) else value
            for value in build_totals(records)
        ])

    # BOM so Excel opens UTF-8 CSV with the right encoding on Windows.
    response = HttpResponse(
        '\ufeff' + buffer.getvalue(),
        content_type='text/csv; charset=utf-8',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{export_filename(year, month, "csv")}"'
    )
    return response


# ── PDF ──────────────────────────────────────────────────────────────────────

def export_pdf(records, year, month) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    records = list(records)
    navy = colors.HexColor('#0f3d73')
    grey = colors.HexColor('#475569')
    line = colors.HexColor('#cbd5e1')
    light = colors.HexColor('#f1f5f9')

    base = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'PayrollTitle', parent=base['Normal'], fontName='Helvetica-Bold',
        fontSize=17, textColor=navy, leading=21, spaceAfter=3,
    )
    sub_style = ParagraphStyle(
        'PayrollSub', parent=base['Normal'], fontName='Helvetica',
        fontSize=9, textColor=grey, leading=12, spaceAfter=12,
    )
    cell_style = ParagraphStyle(
        'PayrollCell', parent=base['Normal'], fontName='Helvetica',
        fontSize=7.5, leading=9.5,
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=f'Payroll Register {month:02d}-{year}',
    )

    story = [
        Paragraph(f'Payroll Register — {period_label(year, month)}', title_style),
        Paragraph(
            f'{len(records)} record(s) · Generated '
            f'{timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")}',
            sub_style,
        ),
    ]

    headers = [COLUMNS[i][0] for i in PDF_COLUMNS]
    data = [[Paragraph(f'<b>{h}</b>', cell_style) for h in headers]]

    def fmt(value):
        return f'{value:,.2f}' if isinstance(value, Decimal) else str(value)

    for row in build_rows(records):
        data.append([Paragraph(fmt(row[i]), cell_style) for i in PDF_COLUMNS])

    if records:
        totals = build_totals(records)
        data.append([
            Paragraph(f'<b>{fmt(totals[i])}</b>' if totals[i] != '' else '', cell_style)
            for i in PDF_COLUMNS
        ])

    if not records:
        data.append([Paragraph('No payroll records for this period.', cell_style)]
                    + [''] * (len(headers) - 1))

    table = Table(data, repeatRows=1, hAlign='LEFT')
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), navy),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, line),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2 if records else -1),
         [colors.white, colors.HexColor('#f8fafc')]),
    ]
    if records:
        style += [
            ('BACKGROUND', (0, -1), (-1, -1), light),
            ('LINEABOVE', (0, -1), (-1, -1), 1, navy),
        ]
    table.setStyle(TableStyle(style))

    story += [table, Spacer(1, 8 * mm)]
    doc.build(story)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="{export_filename(year, month, "pdf")}"'
    )
    return response


# ── View ─────────────────────────────────────────────────────────────────────

class PayrollExportView(LoginRequiredMixin, View):
    """
    Download the payroll register for a period.

    Query params:
        year    — defaults to the current year
        month   — defaults to the current month
        format  — xlsx (default) | csv | pdf
        status  — optional: draft | approved | paid | cancelled | all
    """

    def get(self, request):
        if not can_run_payroll(request.user):
            return HttpResponseForbidden(
                'You do not have permission to download payroll.'
            )

        now = timezone.now()
        try:
            year = int(request.GET.get('year') or now.year)
        except (TypeError, ValueError):
            year = now.year
        try:
            month = int(request.GET.get('month') or now.month)
        except (TypeError, ValueError):
            month = now.month
        if not 1 <= month <= 12:
            month = now.month

        fmt = (request.GET.get('format') or 'xlsx').strip().lower()
        status = (request.GET.get('status') or '').strip().lower()

        records = list(payroll_queryset(year, month, status))

        if fmt == 'csv':
            return export_csv(records, year, month)
        if fmt == 'pdf':
            try:
                return export_pdf(records, year, month)
            except ImportError:
                # reportlab missing — don't 500, give them the data anyway.
                return export_csv(records, year, month)
        try:
            return export_xlsx(records, year, month)
        except ImportError:
            return export_csv(records, year, month)
