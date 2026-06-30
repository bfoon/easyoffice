"""
apps/inventory/importers.py
───────────────────────────
Bulk product import from an Odoo "Product (product.template)" Excel export.

Odoo's standard export (the *.xlsx you get from the Products list view via
Action → Export) carries these columns:

    Currency · Activity State · Product Category · Favorite · Name ·
    Internal Reference · # Product Variants · Sales Price ·
    Quantity On Hand · Unit · Image 128 · Last Updated on ·
    Show On Hand Qty Status Button

This module maps that shape onto EasyOffice's inventory models and seeds
opening stock through the normal services layer (so every quantity lands
as a proper StockMovement, not a silent write).

Design notes
════════════
  • Idempotent on SKU. Re-running with the same file updates existing
    products rather than duplicating them (match on sku, then on name).
  • Odoo "Internal Reference" → our `sku`. It is NOT unique in Odoo and is
    often blank, so we clean + de-duplicate and synthesise a stable SKU
    when missing.
  • "Product Category" is a slash path ("Goods/Toners/HP"). We collapse it
    to our two-level Category (parent → child), auto-creating as needed.
  • Opening stock is applied via services.receive_stock() (positive) or
    services.adjust_stock() (negative — Odoo allows negative on-hand and so
    do we). Skip with seed_stock=False to import the catalog only.
  • "Image 128" in standard exports is usually Odoo's grey placeholder SVG.
    We decode it but skip placeholders by default; pass import_images=True
    to keep whatever is there.

Entry point
═══════════
    from apps.inventory.importers import import_odoo_products
    result = import_odoo_products(
        path_or_filelike,
        location=warehouse,        # where opening stock lands (required if seed_stock)
        actor=request.user,
        seed_stock=True,
        dry_run=False,
    )
    print(result.as_dict())
"""
from __future__ import annotations

import base64
import logging
import re
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Optional

from django.core.files.base import ContentFile
from django.db import transaction
from django.utils.text import slugify

from . import services
from .models import Category, Location, Product

logger = logging.getLogger(__name__)


# ── Expected Odoo headers (we match case-insensitively / loosely) ───────────

COL_CURRENCY   = 'currency'
COL_CATEGORY   = 'product category'
COL_NAME       = 'name'
COL_REF        = 'internal reference'
COL_PRICE      = 'sales price'
COL_QTY        = 'quantity on hand'
COL_UNIT       = 'unit'
COL_IMAGE      = 'image 128'

# Unit-of-measure text from Odoo → our short unit_label
UNIT_MAP = {
    'units':  'unit',
    'unit':   'unit',
    'each':   'unit',
    'pcs':    'unit',
    'kg':     'kg',
    'g':      'g',
    'l':      'litre',
    'm':      'metre',
    'km':     'km',
    'hours':  'hour',
    'hour':   'hour',
    'days':   'day',
}

SKU_MAXLEN  = 40    # matches Product.sku max_length
NAME_MAXLEN = 200   # matches Product.name max_length
CAT_MAXLEN  = 100   # matches Category.name max_length


# ════════════════════════════════════════════════════════════════════════════
# Result container
# ════════════════════════════════════════════════════════════════════════════

@dataclass
class RowIssue:
    row: int
    name: str
    message: str


@dataclass
class ImportResult:
    total_rows:        int = 0
    created:           int = 0
    updated:           int = 0
    skipped:           int = 0
    categories_made:   int = 0
    stock_seeded:      int = 0
    images_imported:   int = 0
    dry_run:           bool = False
    issues:            list = field(default_factory=list)

    def add_issue(self, row: int, name: str, message: str):
        self.issues.append(RowIssue(row=row, name=name, message=message))

    def as_dict(self) -> dict:
        return {
            'total_rows':      self.total_rows,
            'created':         self.created,
            'updated':         self.updated,
            'skipped':         self.skipped,
            'categories_made': self.categories_made,
            'stock_seeded':    self.stock_seeded,
            'images_imported': self.images_imported,
            'dry_run':         self.dry_run,
            'issues':          [
                {'row': i.row, 'name': i.name, 'message': i.message}
                for i in self.issues
            ],
        }


# ════════════════════════════════════════════════════════════════════════════
# Small parsing helpers
# ════════════════════════════════════════════════════════════════════════════

def _clean(value) -> str:
    """Trim, normalise whitespace, treat NaN/None as ''. """
    if value is None:
        return ''
    s = str(value)
    if s.strip().lower() in ('nan', 'none', 'false', '#n/a'):
        # Odoo sometimes emits literal FALSE in boolean-ish text cells;
        # only blank it out for the columns where that's noise.
        if s.strip().lower() in ('nan', 'none'):
            return ''
    s = unicodedata.normalize('NFKC', s).strip()
    return re.sub(r'\s+', ' ', s)


def _to_decimal(value, default=Decimal('0.00')) -> Decimal:
    s = _clean(value).replace(',', '')
    if not s:
        return default
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return default


def _norm_unit(value) -> str:
    s = _clean(value).lower()
    return UNIT_MAP.get(s, s or 'unit')[:20]


def _slug_sku(name: str, taken: set[str]) -> str:
    """Build a stable, unique uppercase SKU from a product name."""
    base = slugify(name).upper().replace('-', '')[:SKU_MAXLEN - 4] or 'ITEM'
    candidate = base
    n = 1
    while candidate in taken or Product.objects.filter(sku__iexact=candidate).exists():
        suffix = f'-{n}'
        candidate = (base[:SKU_MAXLEN - len(suffix)] + suffix)
        n += 1
    return candidate[:SKU_MAXLEN]


def _resolve_columns(raw_headers: list[str]) -> dict[str, str]:
    """Map our canonical lower-case keys to the actual header strings present."""
    lookup = {str(h).strip().lower(): h for h in raw_headers}
    wanted = {
        COL_CURRENCY: COL_CURRENCY,
        COL_CATEGORY: COL_CATEGORY,
        COL_NAME:     COL_NAME,
        COL_REF:      COL_REF,
        COL_PRICE:    COL_PRICE,
        COL_QTY:      COL_QTY,
        COL_UNIT:     COL_UNIT,
        COL_IMAGE:    COL_IMAGE,
    }
    resolved = {}
    for key in wanted:
        if key in lookup:
            resolved[key] = lookup[key]
    return resolved


# ════════════════════════════════════════════════════════════════════════════
# Category path → two-level Category
# ════════════════════════════════════════════════════════════════════════════

def _get_or_create_category(path: str, result: ImportResult, dry_run: bool):
    """
    Odoo category is a slash path, e.g. "Goods/Toners/HP 80A".
    We map it onto our two-level (parent → child) model:
      • 1 segment  → top-level category
      • 2 segments → parent → child
      • 3+ segments→ first segment is parent, the REST joined as the child
                     (keeps the leaf meaningful without exploding the tree)
    Returns a Category instance or None.
    """
    path = _clean(path)
    if not path or path == 'False':
        return None

    segments = [s.strip() for s in path.split('/') if s.strip()]
    if not segments:
        return None

    if len(segments) == 1:
        parent_name, child_name = None, segments[0]
    else:
        parent_name = segments[0]
        child_name = ' / '.join(segments[1:])

    # Category.name is capped at 100 chars.
    if parent_name:
        parent_name = parent_name[:CAT_MAXLEN]
    child_name = child_name[:CAT_MAXLEN]

    parent = None
    if parent_name:
        if dry_run:
            parent = Category.objects.filter(parent__isnull=True, name__iexact=parent_name).first()
            if parent is None:
                result.categories_made += 1
                return None  # can't attach a child to a not-yet-created parent in dry-run
        else:
            parent, made = Category.objects.get_or_create(
                parent__isnull=True, name=parent_name,
                defaults={'name': parent_name},
            )
            if made:
                result.categories_made += 1

    if dry_run:
        existing = Category.objects.filter(
            parent=parent, name__iexact=child_name,
        ).first()
        if existing is None:
            result.categories_made += 1
        return existing

    cat, made = Category.objects.get_or_create(
        parent=parent, name=child_name,
    )
    if made:
        result.categories_made += 1
    return cat


# ════════════════════════════════════════════════════════════════════════════
# Image decode
# ════════════════════════════════════════════════════════════════════════════

_PLACEHOLDER_MARKERS = (b'<svg', b'placeholder')

def _decode_image(b64: str):
    """
    Returns (ContentFile, ext) or None. Skips Odoo's grey placeholder SVGs.
    """
    b64 = _clean(b64)
    if not b64:
        return None
    try:
        raw = base64.b64decode(b64)
    except Exception:
        return None
    head = raw[:64].lower()
    # Odoo's "Image 128" placeholder is an SVG box — not a real photo.
    if head.startswith(b'<svg'):
        return None
    # Detect common raster signatures
    if raw[:3] == b'\xff\xd8\xff':
        ext = 'jpg'
    elif raw[:8] == b'\x89PNG\r\n\x1a\n':
        ext = 'png'
    elif raw[:6] in (b'GIF87a', b'GIF89a'):
        ext = 'gif'
    elif raw[:4] == b'RIFF' and raw[8:12] == b'WEBP':
        ext = 'webp'
    else:
        return None
    return ContentFile(raw), ext


# ════════════════════════════════════════════════════════════════════════════
# Workbook reader (openpyxl — no pandas dependency)
# ════════════════════════════════════════════════════════════════════════════

def _read_xlsx(path_or_filelike, sheet_name=0):
    """
    Read the first row as headers and return (headers, rows) where each row
    is a dict keyed by the header string. Uses openpyxl in read-only mode so
    large exports stream cheaply.

    `sheet_name` may be an int index (0 = first sheet) or a sheet title.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path_or_filelike, read_only=True, data_only=True)
    try:
        if isinstance(sheet_name, int):
            ws = wb[wb.sheetnames[sheet_name]]
        else:
            ws = wb[sheet_name]

        row_iter = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(row_iter)
        except StopIteration:
            return [], []

        headers = [('' if h is None else str(h)) for h in raw_headers]

        rows = []
        for values in row_iter:
            # Skip fully-blank rows (trailing empties are common in exports)
            if values is None or all(v is None or str(v).strip() == '' for v in values):
                continue
            row = {}
            for i, h in enumerate(headers):
                row[h] = values[i] if i < len(values) else None
            rows.append(row)
        return headers, rows
    finally:
        wb.close()


# ════════════════════════════════════════════════════════════════════════════
# Main entry point
# ════════════════════════════════════════════════════════════════════════════

def import_odoo_products(
    path_or_filelike,
    *,
    location: Optional[Location] = None,
    actor=None,
    seed_stock: bool = True,
    import_images: bool = False,
    default_currency: str = 'GMD',
    update_existing: bool = True,
    sheet_name=0,
    dry_run: bool = False,
) -> ImportResult:
    """
    Parse an Odoo product export and create/update inventory products.

    Parameters
    ----------
    path_or_filelike : str | Path | file-like
        The .xlsx file (path or an uploaded file object).
    location : Location
        Where opening stock is recorded. Required when seed_stock=True.
    actor : User
        Recorded on stock movements and as created_by.
    seed_stock : bool
        Apply "Quantity On Hand" as opening stock. Off → catalog only.
    import_images : bool
        Keep decoded raster images (placeholders are always skipped).
    update_existing : bool
        If a product with the same SKU/name exists, update it. Else skip.
    dry_run : bool
        Parse and validate but write nothing. Returns a full report.
    """
    from openpyxl import load_workbook  # local import keeps openpyxl optional at app load

    if seed_stock and location is None and not dry_run:
        raise ValueError('A location is required when seed_stock=True.')

    headers, data_rows = _read_xlsx(path_or_filelike, sheet_name)
    cols = _resolve_columns(headers)

    if COL_NAME not in cols:
        raise ValueError(
            "This doesn't look like an Odoo product export — no 'Name' column found. "
            f"Columns seen: {headers}"
        )

    result = ImportResult(total_rows=len(data_rows), dry_run=dry_run)
    taken_skus: set[str] = set()

    # We run the whole thing in a transaction; on dry_run we roll back at the end.
    transaction.set_autocommit(False)
    try:
        for idx, row in enumerate(data_rows):
            excel_row = idx + 2  # header is row 1
            full_name = _clean(row.get(cols.get(COL_NAME)))
            if not full_name:
                result.skipped += 1
                result.add_issue(excel_row, '(blank)', 'No product name — skipped.')
                continue

            # Product.name is capped at 200 chars. Long Odoo names are really
            # full spec sheets — keep the whole thing in `description` and use
            # a trimmed-on-a-word-boundary version as the name.
            name = full_name
            overflow_description = ''
            if len(full_name) > NAME_MAXLEN:
                cut = full_name[:NAME_MAXLEN].rsplit(' ', 1)[0] or full_name[:NAME_MAXLEN]
                name = (cut[:NAME_MAXLEN - 1].rstrip() + '…')
                overflow_description = full_name
                result.add_issue(
                    excel_row, name,
                    f'Name was {len(full_name)} chars — trimmed to fit; '
                    f'full text saved to description.',
                )

            # ── SKU resolution ────────────────────────────────────────────
            raw_ref = _clean(row.get(cols.get(COL_REF), '')) if COL_REF in cols else ''
            sku = raw_ref.upper()[:SKU_MAXLEN]
            if not sku:
                sku = _slug_sku(name, taken_skus)
            elif sku in taken_skus or (
                Product.objects.filter(sku__iexact=sku)
                .exclude(name__iexact=name).exists()
            ):
                # Collision with a DIFFERENT product → make it unique.
                sku = _slug_sku(f'{name}', taken_skus)
            taken_skus.add(sku)

            # ── Field values ──────────────────────────────────────────────
            sell_price = _to_decimal(row.get(cols.get(COL_PRICE))) if COL_PRICE in cols else Decimal('0.00')
            unit_label = _norm_unit(row.get(cols.get(COL_UNIT))) if COL_UNIT in cols else 'unit'
            currency = (_clean(row.get(cols.get(COL_CURRENCY))) or default_currency)[:3] if COL_CURRENCY in cols else default_currency
            category = _get_or_create_category(
                row.get(cols.get(COL_CATEGORY), ''), result, dry_run,
            ) if COL_CATEGORY in cols else None
            qty = _to_decimal(row.get(cols.get(COL_QTY))) if COL_QTY in cols else Decimal('0.00')

            # Heuristic: zero price + an Hours/service unit → SERVICE kind.
            kind = Product.Kind.STOCKED
            if unit_label in ('hour', 'day') and sell_price >= 0:
                kind = Product.Kind.SERVICE

            # ── Find existing (sku first, then exact name) ────────────────
            existing = (
                Product.objects.filter(sku__iexact=sku).first()
                or Product.objects.filter(name__iexact=name).first()
            )

            if existing and not update_existing:
                result.skipped += 1
                result.add_issue(excel_row, name, 'Already exists — skipped (update disabled).')
                continue

            defaults = dict(
                name=name,
                sell_price=sell_price,
                unit_label=unit_label,
                currency=currency,
                category=category,
                kind=kind,
            )
            # Only set description when the name overflowed, so we don't
            # blank out a description the user may have written by hand.
            if overflow_description:
                defaults['description'] = overflow_description

            if existing:
                for f_, v_ in defaults.items():
                    setattr(existing, f_, v_)
                # don't clobber a manually-set sku on update
                product = existing
                product.save()
                result.updated += 1
            else:
                product = Product.objects.create(
                    sku=sku,
                    created_by=actor,
                    is_for_internal=(kind == Product.Kind.STOCKED and sell_price == 0),
                    **defaults,
                )
                result.created += 1

            # ── Image (optional) ──────────────────────────────────────────
            if import_images and COL_IMAGE in cols:
                decoded = _decode_image(row.get(cols.get(COL_IMAGE)))
                if decoded:
                    content, ext = decoded
                    if not dry_run:
                        product.image.save(f'{slugify(sku) or "product"}.{ext}', content, save=True)
                    result.images_imported += 1

            # ── Opening stock ─────────────────────────────────────────────
            if seed_stock and kind == Product.Kind.STOCKED and qty != 0:
                if dry_run or location is None:
                    result.stock_seeded += 1
                else:
                    try:
                        if qty > 0:
                            services.receive_stock(
                                product=product, location=location, quantity=qty,
                                actor=actor, source_kind='odoo_import',
                                source_ref=f'row:{excel_row}',
                                notes='Opening balance — Odoo import',
                            )
                        else:
                            services.adjust_stock(
                                product=product, location=location, delta=qty,
                                actor=actor, source_kind='odoo_import',
                                source_ref=f'row:{excel_row}',
                                notes='Opening balance (negative) — Odoo import',
                            )
                        result.stock_seeded += 1
                    except Exception as exc:  # noqa: BLE001
                        result.add_issue(excel_row, name, f'Stock seed failed: {exc}')

        if dry_run:
            transaction.rollback()
        else:
            transaction.commit()
    except Exception:
        transaction.rollback()
        raise
    finally:
        transaction.set_autocommit(True)

    return result