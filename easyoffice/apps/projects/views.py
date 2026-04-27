from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, DetailView, View
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q, Count
from django.utils import timezone
from .models import (
    Project, Milestone, ProjectUpdate, Risk,
    Survey, SurveyQuestion, SurveyResponse, SurveyAnswer,
    TrackingSheet, TrackingRow, ProjectLocation, LocationImport,
)
from apps.messaging.models import ChatRoom
import csv
import json
from collections import Counter
from apps.core.models import User
from django.db.models import Sum
from apps.finance.models import PurchaseRequest, Payment, EmployeeFinanceRequest
from apps.files.models import SharedFile
import io
import requests
import logging
from datetime import date, timedelta, datetime, time
from django.core.files.base import ContentFile
from django.db import transaction
from django.views.generic import ListView
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_GET
from apps.projects.geo_utils import (
    parse_geo_answer, lookup_ip_location, reverse_geocode,
    plus_code_to_latlng, is_plus_code, geo_answer_to_columns,
)
import os
from django.urls import reverse

def _is_project_member(user, project):
    return (
        user.is_superuser or
        project.project_manager == user or
        project.team_members.filter(id=user.id).exists()
    )

def _is_project_owner(user, project):
    return user.is_superuser or project.project_manager_id == user.id


def _milestone_due_datetime(due_date):
    if not due_date:
        return None

    # Accept either a date / datetime instance or an ISO 'YYYY-MM-DD' string.
    if isinstance(due_date, str):
        try:
            due_date = datetime.strptime(due_date.strip(), '%Y-%m-%d').date()
        except ValueError:
            return None

    # If a datetime was passed, pull just the date out so we can stamp 17:00.
    if isinstance(due_date, datetime):
        due_date = due_date.date()

    dt = datetime.combine(due_date, time(hour=17, minute=0))

    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_current_timezone())

    return dt


def _create_or_update_milestone_task(milestone, actor):
    """
    Create or sync the task linked to a project milestone.
    Selecting an assignee on a milestone automatically gives that person
    a task with the same project and deadline.
    """
    if not milestone.assigned_to_id:
        return None

    from apps.tasks.models import Task
    from apps.core.models import CoreNotification

    project = milestone.project
    due_dt = _milestone_due_datetime(milestone.due_date)

    priority = project.priority
    if priority not in dict(Task.Priority.choices):
        priority = Task.Priority.MEDIUM

    title = f'{project.code}: {milestone.name}'
    description = (
        f'Project milestone task for: {project.name}\n\n'
        f'Milestone: {milestone.name}\n\n'
        f'{milestone.description or ""}'
    ).strip()

    if milestone.task_id:
        task = milestone.task
        task.title = title
        task.description = description
        task.assigned_to = milestone.assigned_to
        task.project = project
        task.due_date = due_dt
        task.priority = priority

        if milestone.status == Milestone.Status.COMPLETED:
            task.status = Task.Status.DONE
            task.progress_pct = 100
            task.completed_at = timezone.now()
        elif milestone.status == Milestone.Status.IN_PROGRESS and task.status == Task.Status.TODO:
            task.status = Task.Status.IN_PROGRESS

        task.save(update_fields=[
            'title',
            'description',
            'assigned_to',
            'project',
            'due_date',
            'priority',
            'status',
            'progress_pct',
            'completed_at',
            'updated_at',
        ])
    else:
        task = Task.objects.create(
            title=title,
            description=description,
            assigned_to=milestone.assigned_to,
            assigned_by=actor,
            project=project,
            priority=priority,
            status=Task.Status.TODO,
            due_date=due_dt,
        )
        milestone.task = task
        milestone.save(update_fields=['task', 'updated_at'])

    CoreNotification.objects.create(
        recipient=milestone.assigned_to,
        sender=actor,
        notification_type='project',
        title='Project Milestone Assigned',
        message=f'You have been assigned milestone "{milestone.name}" under project "{project.name}".',
        link=reverse('project_detail', kwargs={'pk': project.pk}),
    )

    return task


def _log_milestone_activity(milestone, author, content, *, title='', is_milestone_update=True):
    """
    Create a ProjectUpdate row tied to this milestone so the milestone's
    progress (assignment, completion, manual notes) appears in the
    project's Activity Log. Idempotent-ish: callers should pass distinct
    `content` per event — we don't try to dedupe identical posts.
    """
    if not milestone or not author:
        return None

    project = milestone.project
    return ProjectUpdate.objects.create(
        project=project,
        author=author,
        title=(title or f'Milestone — {milestone.name}')[:200],
        content=content,
        progress_at_time=project.progress_pct,
        status_at_time=project.status,
        is_milestone_update=is_milestone_update,
        milestone=milestone,
    )


def sync_milestone_from_task(task):
    """
    Called from apps.tasks.views.TaskQuickUpdateView whenever a task is
    saved. If the task is the auto-created counterpart of a Milestone:
      * mark the milestone Completed when the task is Done, and
      * post a ProjectUpdate tying the latest completion-note comment to
        the project's Activity Log so supervisors can see what was done.

    Lives in apps.projects.views (rather than tasks) because it crosses
    app boundaries and the milestone-side ownership lives here.
    """
    milestone = getattr(task, 'project_milestone', None)
    if milestone is None:
        return

    changed = []
    if task.status == 'done' and milestone.status != Milestone.Status.COMPLETED:
        milestone.status = Milestone.Status.COMPLETED
        milestone.completed_at = task.completed_at or timezone.now()
        changed += ['status', 'completed_at', 'updated_at']
    elif task.status != 'done' and milestone.status == Milestone.Status.COMPLETED:
        # Task got reopened — un-complete the milestone too.
        milestone.status = Milestone.Status.IN_PROGRESS
        milestone.completed_at = None
        changed += ['status', 'completed_at', 'updated_at']

    if changed:
        milestone.save(update_fields=list(set(changed)))

        # If this transition was task-done, log it to the project's
        # Activity Log with the latest completion-note comment.
        if task.status == 'done':
            from apps.tasks.models import TaskComment
            note = (
                TaskComment.objects
                .filter(task=task, content__startswith='✅ Completion note:')
                .order_by('-created_at')
                .first()
            )
            content_lines = [f'Milestone "{milestone.name}" was completed.']
            if note:
                # Strip the "✅ Completion note:\n" prefix the task layer adds.
                stripped = note.content.split('\n', 1)
                completion_text = stripped[1].strip() if len(stripped) > 1 else ''
                if completion_text:
                    content_lines.append('')
                    content_lines.append(completion_text)
            actor = (note.author if note else task.assigned_to) or task.assigned_by
            _log_milestone_activity(
                milestone,
                actor,
                '\n'.join(content_lines),
                title=f'Milestone completed — {milestone.name}',
            )


logger = logging.getLogger(__name__)

# --------------------------Helper ---------------------------------

# Column-header aliases used to auto-detect which source column feeds which
# target field. Lower-cased, stripped. Order in each list = preference.
_LOCATION_FIELD_ALIASES = {
    'name': ['name', 'site', 'site name', 'location', 'title', 'label'],
    'latitude': ['latitude', 'lat', 'y', 'geo_lat', 'coord_lat'],
    'longitude': ['longitude', 'lng', 'lon', 'long', 'x', 'geo_lng', 'geo_lon'],
    'address': ['address', 'street', 'location_address', 'full_address'],
    'description': ['description', 'desc', 'notes', 'details', 'remarks'],
    'category': ['category', 'type', 'kind', 'group'],
    'status': ['status', 'state'],
}

# Survey geo answer only needs lat/lng/address/accuracy/source
_SURVEY_GEO_ALIASES = {
    'latitude': _LOCATION_FIELD_ALIASES['latitude'],
    'longitude': _LOCATION_FIELD_ALIASES['longitude'],
    'address': _LOCATION_FIELD_ALIASES['address'],
    'accuracy_m': ['accuracy', 'accuracy_m', 'accuracy_meters', 'acc', 'precision_m'],
    'source': ['source', 'origin', 'captured_by', 'method'],
    'name': _LOCATION_FIELD_ALIASES['name'],  # stored into respondent_name
}


def _auto_detect_mapping(headers, aliases):
    """Given a list of source column headers and a target alias map, return
    {target_field: best_matching_header_or_''} for every target field."""
    seen = set()
    mapping = {}
    lower_headers = {h.strip().lower(): h for h in headers if h}
    for target, alist in aliases.items():
        picked = ''
        for alias in alist:
            if alias in lower_headers:
                candidate = lower_headers[alias]
                if candidate not in seen:
                    picked = candidate
                    seen.add(candidate)
                    break
        mapping[target] = picked
    return mapping


def _read_tabular(file_handle, filename):
    """Return (headers, rows) from an uploaded CSV or XLSX. Rows are list of dicts."""
    fname = (filename or '').lower()
    if fname.endswith('.csv'):
        text = file_handle.read().decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        headers = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader]
        return headers, rows
    if fname.endswith(('.xlsx', '.xls')):
        import openpyxl
        wb = openpyxl.load_workbook(file_handle, data_only=True)
        ws = wb.active
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = [str(h).strip() if h is not None else '' for h in header_cells]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and str(v).strip() for v in row):
                continue
            rows.append(dict(zip(headers, [
                '' if v is None else str(v).strip() for v in row
            ])))
        return headers, rows
    raise ValueError('Unsupported file type. Use .csv, .xlsx or .xls.')


def _parse_latlng(val, ref_lat=13.4549, ref_lng=-16.5790):
    """Coerce a cell value into a float or None.
    Handles plain floats, degree symbols, and Google Plus Codes (full & short).
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if is_plus_code(s):
        lat, _lng = plus_code_to_latlng(s, ref_lat, ref_lng)
        return lat
    s = s.replace(',', '.').replace('°', '')
    try:
        return float(s)
    except ValueError:
        return None


def _resolve_latlng(lat_raw, lng_raw, ref_lat=13.4549, ref_lng=-16.5790):
    """
    Convert a pair of raw cell values to (lat_float, lng_float) or (None, None).
    Handles Plus Codes in the latitude OR longitude column.
    """
    lat_s = str(lat_raw or '').strip()
    lng_s = str(lng_raw or '').strip()
    if lat_s and is_plus_code(lat_s):
        return plus_code_to_latlng(lat_s, ref_lat, ref_lng)
    if lng_s and is_plus_code(lng_s):
        return plus_code_to_latlng(lng_s, ref_lat, ref_lng)
    try:
        return (
            float(lat_s.replace(',', '.').replace('°', '')),
            float(lng_s.replace(',', '.').replace('°', '')),
        )
    except (ValueError, TypeError):
        return None, None


class OSMTileProxyView(View):
    """
    Proxy OSM raster tiles through Django so the browser never contacts
    the tile server directly.  Direct browser requests to OSM volunteer
    servers are blocked (HTTP 403 "Referer required") — proxying through
    Django sets the correct User-Agent and Referer headers.

    Tiles are cached in Django's cache backend for 24 h to avoid
    hammering OSM's servers on repeated map loads.

    URL (already in urls.py):
        surveys/osm-tiles/<int:z>/<int:x>/<int:y>.png
    """

    TILE_TIMEOUT = 10
    SUBDOMAINS   = ['a', 'b', 'c']
    USER_AGENT   = 'EasyOffice/1.0 (internal map proxy; contact admin@easyoffice.local)'

    def get(self, request, z, x, y):
        import hashlib
        from django.core.cache import cache

        sub      = self.SUBDOMAINS[(x + y) % 3]
        tile_url = f'https://{sub}.tile.openstreetmap.org/{z}/{x}/{y}.png'
        cache_key = 'osm_tile:' + hashlib.md5(tile_url.encode()).hexdigest()

        cached = cache.get(cache_key)
        if cached:
            resp = HttpResponse(cached, content_type='image/png')
            resp['Cache-Control'] = 'public, max-age=86400'
            return resp

        try:
            r = requests.get(
                tile_url,
                headers={
                    'User-Agent': self.USER_AGENT,
                    'Referer':    request.build_absolute_uri('/'),
                },
                timeout=self.TILE_TIMEOUT,
            )
        except requests.RequestException:
            return HttpResponse(status=502)

        if r.status_code != 200:
            return HttpResponse(status=r.status_code)

        tile_bytes = r.content
        cache.set(cache_key, tile_bytes, 86400)

        resp = HttpResponse(tile_bytes, content_type=r.headers.get('Content-Type', 'image/png'))
        resp['Cache-Control'] = 'public, max-age=86400'
        return resp


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _read_spreadsheet(file_obj=None, shared_file=None):
    """
    Read a CSV or XLSX file.  Supply exactly one of file_obj (InMemoryUploadedFile)
    or shared_file (SharedFile model instance).
    Returns (headers: list[str], rows: list[dict], error: str|None).
    """
    try:
        if shared_file:
            name = shared_file.name.lower()
            data = shared_file.file.read()
        else:
            name = file_obj.name.lower()
            data = file_obj.read()

        if name.endswith('.csv'):
            text = data.decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
            headers = list(reader.fieldnames or [])
            if not headers and rows:
                headers = list(rows[0].keys())
            return headers, rows, None

        elif name.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return [], [], 'Empty spreadsheet'
            headers = [str(c or '').strip() for c in all_rows[0]]
            rows = []
            for raw in all_rows[1:]:
                row = {headers[i]: (str(v).strip() if v is not None else '')
                       for i, v in enumerate(raw) if i < len(headers)}
                rows.append(row)
            return headers, rows, None

        else:
            return [], [], 'Unsupported file type. Use CSV or XLSX.'

    except Exception as e:
        logger.exception('_read_spreadsheet error')
        return [], [], str(e)


def _auto_suggest_mapping(headers: list, target: str) -> dict:
    """
    Guess column mapping from header names.
    Returns {field_key: header_name}.
    """
    norm = {h.lower().replace(' ', '_').replace('-', '_'): h for h in headers}
    mapping = {}

    def pick(*candidates):
        for c in candidates:
            if c in norm:
                return norm[c]
        return ''

    if target == 'locations':
        mapping['name'] = pick('name', 'location_name', 'site', 'site_name', 'title')
        mapping['latitude'] = pick('latitude', 'lat', 'y', 'plus_code', 'pluscode', 'olc',
                                   'open_location_code', 'google_plus_code')
        mapping['longitude'] = pick('longitude', 'lng', 'lon', 'long', 'x')
        mapping['address'] = pick('address', 'addr', 'location')
        mapping['description'] = pick('description', 'desc', 'details', 'notes')
        mapping['category'] = pick('category', 'cat', 'type')
        mapping['status'] = pick('status', 'state')
        mapping['notes'] = pick('notes', 'remarks', 'comment', 'comments')
    else:  # survey geo
        mapping['latitude'] = pick('latitude', 'lat', 'y', 'plus_code', 'pluscode', 'olc')
        mapping['longitude'] = pick('longitude', 'lng', 'lon', 'long', 'x')
        mapping['address'] = pick('address', 'addr')
        mapping['accuracy_m'] = pick('accuracy_m', 'accuracy', 'acc')
        mapping['source'] = pick('source', 'src', 'method')
        mapping['name'] = pick('name', 'respondent', 'respondent_name')

    return {k: v for k, v in mapping.items() if v}


# ─────────────────────────────────────────────────────────────────────────────
# 1. Enhanced Location Map Export — CSV (streaming) + new "Save to Files"
# ─────────────────────────────────────────────────────────────────────────────

class SurveyExportView(LoginRequiredMixin, View):
    """Combined endpoint: `?format=csv|xlsx&save_to_files=0|1`."""

    def get(self, request, pk, sid):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()
        survey = get_object_or_404(Survey, pk=sid, project=project)

        fmt = (request.GET.get('format') or 'csv').lower()
        content, out_name, mime = _build_survey_export(survey, fmt=fmt)

        if request.GET.get('save_to_files') == '1':
            sf = SharedFile.objects.create(
                name=out_name,
                file=ContentFile(content, name=out_name),
                project=project,
                uploaded_by=request.user,
                visibility='private',
                description=f'Survey export · {survey.title}',
                tags='export,survey',
                file_size=len(content),
                file_type=mime.split(';')[0].strip(),
            )
            try:
                sf.file_hash = sf.compute_hash()
                sf.save(update_fields=['file_hash'])
            except Exception:
                pass
            return JsonResponse({
                'ok': True,
                'message': f'Saved "{out_name}" to your Files.',
                'file_id': str(sf.pk),
                'file_name': sf.name,
            })

        resp = HttpResponse(content, content_type=mime)
        resp['Content-Disposition'] = f'attachment; filename="{out_name}"'
        return resp

# ─────────────────────────────────────────────────────────────────────────────
# 3. Import — multi-step flow
#    Step 1: POST /import/preview/       → return headers + sample rows + suggested mapping
#    Step 2: POST /import/commit/        → perform the insert with the chosen mapping
# ─────────────────────────────────────────────────────────────────────────────

def _user_can_access_file(user, shared_file):
    """Minimal access check — user owns the file or it's a project document
    for a project the user is a member of."""
    if shared_file.uploaded_by_id == user.id or user.is_superuser:
        return True
    if shared_file.project_id:
        try:
            return _is_project_member(user, shared_file.project)
        except Exception:
            return False
    # Fall back to visibility rules
    if shared_file.visibility == 'office':
        return True
    if user in shared_file.shared_with.all():
        return True
    return False


def _resolve_import_source(request, project):
    """
    Given a request, return (file_obj_like, filename, shared_file_or_None, filesize).

    The file_obj_like is always a binary readable stream suitable for csv /
    openpyxl — either the uploaded UploadedFile or the SharedFile.file field.

    Accepts EITHER:
        POST 'import_file'           (new upload, multipart)
        POST 'shared_file_id'        (pick from existing SharedFile)
    """
    sf_id = request.POST.get('shared_file_id') or ''
    if sf_id:
        try:
            sf = SharedFile.objects.get(pk=sf_id)
        except SharedFile.DoesNotExist:
            return None, None, None, 0
        if not _user_can_access_file(request.user, sf):
            return None, None, None, 0
        try:
            f = sf.file
            f.open('rb')
            return f, sf.name, sf, int(sf.file_size or 0)
        except Exception:
            return None, None, None, 0

    upload = request.FILES.get('import_file')
    if upload:
        return upload, upload.name, None, int(getattr(upload, 'size', 0) or 0)
    return None, None, None, 0

def _build_survey_export(survey, fmt='csv'):
    """
    Build a survey export as CSV or XLSX.

    Returns:
        content_bytes, output_filename, mime_type
    """
    fmt = (fmt or 'csv').lower().strip()
    if fmt not in ('csv', 'xlsx'):
        fmt = 'csv'

    questions = survey.questions.all().order_by('order')
    responses = survey.responses.select_related('respondent').prefetch_related(
        'answers', 'answers__question'
    ).order_by('-submitted_at')

    header = [
        'Response ID',
        'Respondent Name',
        'Respondent User',
        'Is Public Response',
        'Submitted At',
        'IP Address',
        'Device ID',
        'User Agent',
    ]

    geo_question_ids = set()
    for q in questions:
        if q.q_type == 'geolocation':
            geo_question_ids.add(str(q.pk))
            header.extend([
                f'{q.text} — Latitude',
                f'{q.text} — Longitude',
                f'{q.text} — Accuracy (m)',
                f'{q.text} — Source',
                f'{q.text} — Address',
                f'{q.text} — Captured At',
            ])
        else:
            header.append(q.text)

    rows = []
    for resp in responses:
        row = [
            str(resp.pk),
            resp.respondent_name or '',
            str(resp.respondent) if resp.respondent else '',
            'Yes' if resp.is_public_response else 'No',
            resp.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if resp.submitted_at else '',
            resp.ip_address or '',
            resp.device_id or '',
            resp.user_agent or '',
        ]

        answers_by_question = {
            str(ans.question_id): ans
            for ans in resp.answers.all()
        }

        for q in questions:
            ans = answers_by_question.get(str(q.pk))

            if not ans:
                if str(q.pk) in geo_question_ids:
                    row.extend(['', '', '', '', '', ''])
                else:
                    row.append('')
                continue

            raw_value = ans.value or ''

            if str(q.pk) in geo_question_ids:
                geo_cols = geo_answer_to_columns(raw_value)
                row.extend([
                    geo_cols.get('latitude', ''),
                    geo_cols.get('longitude', ''),
                    geo_cols.get('accuracy_m', ''),
                    geo_cols.get('source', ''),
                    geo_cols.get('address', ''),
                    geo_cols.get('captured_at', ''),
                ])
            else:
                if q.q_type == 'yes_no':
                    v = str(raw_value).strip().lower()
                    if v in ('true', '1', 'yes', 'y', 'on'):
                        row.append('Yes')
                    elif v in ('false', '0', 'no', 'n', 'off'):
                        row.append('No')
                    else:
                        row.append(raw_value)
                else:
                    row.append(raw_value)

        rows.append(row)

    safe_title = (survey.title or 'survey').strip().replace(' ', '_')

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Survey Export'

        ws.append(header)
        for r in rows:
            ws.append(r)

        header_fill = PatternFill(fill_type='solid', fgColor='1E3A5F')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = '' if cell.value is None else str(cell.value)
                except Exception:
                    val = ''
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

        ws.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return (
            output.getvalue(),
            f'{safe_title}_export.xlsx',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # default csv
    import io
    import csv

    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(header)
    writer.writerows(rows)

    return (
        stream.getvalue().encode('utf-8-sig'),
        f'{safe_title}_export.csv',
        'text/csv; charset=utf-8'
    )


class LocationImportPreviewView(LoginRequiredMixin, View):

    def post(self, request, pk):
        from apps.projects.models import Project
        project = get_object_or_404(Project, pk=pk)

        target = request.POST.get('target', 'locations')
        upload = request.FILES.get('import_file')
        shared_fid = request.POST.get('shared_file_id', '').strip()

        if upload:
            headers, rows, err = _read_spreadsheet(file_obj=upload)
        elif shared_fid:
            from apps.files.models import SharedFile
            sf = get_object_or_404(SharedFile, pk=shared_fid)
            headers, rows, err = _read_spreadsheet(shared_file=sf)
        else:
            return JsonResponse({'ok': False, 'error': 'No file provided.'})

        if err:
            return JsonResponse({'ok': False, 'error': err})

        suggested = _auto_suggest_mapping(headers, target)

        return JsonResponse({
            'ok': True,
            'headers': headers,
            'rows': rows[:10],
            'total_rows': len(rows),
            'suggested_mapping': suggested,
            'saved_mapping': {},  # extend: load from DB per (project, filename)
        })


class LocationImportCommitView(LoginRequiredMixin, View):
    """
    POST — performs the actual insert.

    Required POST fields:
        target             : 'locations' | 'survey'
        column_mapping_json: JSON {target_field: source_header_or_''}
        shared_file_id     : (optional) UUID of the SharedFile used as source
        import_file        : (optional) UploadedFile
        survey_id          : (when target=survey) UUID of the survey
        question_id        : (when target=survey) UUID of the geolocation question
    """

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return JsonResponse({'error': 'Forbidden'}, status=403)

        target = (request.POST.get('target') or 'locations').strip()

        # Parse mapping (what target field ← what source header)
        try:
            mapping = json.loads(request.POST.get('column_mapping_json') or '{}')
            if not isinstance(mapping, dict):
                raise ValueError('mapping must be object')
        except Exception as e:
            return JsonResponse({'error': f'Bad mapping: {e}'}, status=400)

        file_obj, filename, shared_file, filesize = _resolve_import_source(request, project)
        if file_obj is None:
            return JsonResponse({'error': 'No file provided.'}, status=400)

        try:
            headers, rows = _read_tabular(file_obj, filename)
        except Exception as e:
            return JsonResponse({'error': f'Could not read file: {e}'}, status=400)
        finally:
            try:
                if shared_file:
                    file_obj.close()
            except Exception:
                pass

        # ── Branch on target ─────────────────────────────────────────────────
        if target == 'survey':
            result = self._commit_survey(request, project, rows, mapping,
                                         filename, shared_file, filesize)
        else:
            result = self._commit_locations(request, project, rows, mapping,
                                            filename, shared_file, filesize)
        return JsonResponse(result)

    # ── Project Locations target ─────────────────────────────────────────────
    def _commit_locations(self, request, project, rows, mapping, filename, shared_file, filesize):
        def cell(row, target_field):
            src = mapping.get(target_field) or ''
            return row.get(src, '') if src else ''

        created = 0
        skipped = 0
        skip_notes = []

        with transaction.atomic():
            batch = LocationImport.objects.create(
                project=project,
                target=LocationImport.Target.PROJECT_LOCATIONS,
                source_file=shared_file,
                source_filename=filename or '',
                source_filesize=filesize,
                column_mapping=mapping,
                imported_by=request.user,
            )

            # Reference point for short Plus Code recovery
            _first = project.locations.order_by('created_at').first()
            _ref_lat = float(_first.latitude)  if _first else 13.4549
            _ref_lng = float(_first.longitude) if _first else -16.5790

            for i, row in enumerate(rows, start=2):  # +1 for header, +1 for 1-indexed
                lat, lng = _resolve_latlng(
                    cell(row, 'latitude'), cell(row, 'longitude'), _ref_lat, _ref_lng
                )
                name = (cell(row, 'name') or '').strip() or f'Imported row {i-1}'
                if lat is None or lng is None:
                    skipped += 1
                    if len(skip_notes) < 10:
                        skip_notes.append(f'Row {i}: missing/invalid lat or lng')
                    continue
                if not (-90 <= lat <= 90) or not (-180 <= lng <= 180):
                    skipped += 1
                    if len(skip_notes) < 10:
                        skip_notes.append(f'Row {i}: out-of-range coordinates')
                    continue
                ProjectLocation.objects.create(
                    project=project,
                    name=name[:200],
                    description=(cell(row, 'description') or '')[:2000],
                    address=(cell(row, 'address') or '')[:400],
                    latitude=lat,
                    longitude=lng,
                    category=(cell(row, 'category') or '')[:100],
                    status=(cell(row, 'status') or 'pending').lower()[:20],
                    notes=(cell(row, 'notes') or '')[:2000] if 'notes' in mapping else '',
                    import_batch=batch,
                )
                created += 1

            batch.row_count = created
            batch.skipped_count = skipped
            batch.notes = '\n'.join(skip_notes)
            batch.save(update_fields=['row_count', 'skipped_count', 'notes'])

        return {
            'ok': True,
            'import_id': str(batch.pk),
            'created': created,
            'skipped': skipped,
            'skipped_notes': skip_notes,
            'redirect_url': reverse('project_locations', kwargs={'pk': project.pk}),
        }

    # ── Survey geo answers target ────────────────────────────────────────────
    def _commit_survey(self, request, project, rows, mapping, filename, shared_file, filesize):
        from .models import Survey, SurveyQuestion, SurveyResponse, SurveyAnswer

        survey_id = request.POST.get('survey_id') or ''
        question_id = request.POST.get('question_id') or ''
        if not (survey_id and question_id):
            return {'error': 'survey_id and question_id are required for survey imports.'}

        try:
            survey = Survey.objects.get(pk=survey_id, project=project)
            question = SurveyQuestion.objects.get(pk=question_id, survey=survey, q_type='geolocation')
        except (Survey.DoesNotExist, SurveyQuestion.DoesNotExist):
            return {'error': 'Survey or question not found (must be a geolocation question).'}

        def cell(row, target_field):
            src = mapping.get(target_field) or ''
            return row.get(src, '') if src else ''

        created = 0
        skipped = 0
        skip_notes = []

        with transaction.atomic():
            batch = LocationImport.objects.create(
                project=project,
                target=LocationImport.Target.SURVEY_ANSWERS,
                survey=survey,
                survey_question=question,
                source_file=shared_file,
                source_filename=filename or '',
                source_filesize=filesize,
                column_mapping=mapping,
                imported_by=request.user,
            )

            # Reference point for short Plus Code recovery
            _first = project.locations.order_by('created_at').first()
            _ref_lat = float(_first.latitude)  if _first else 13.4549
            _ref_lng = float(_first.longitude) if _first else -16.5790

            for i, row in enumerate(rows, start=2):
                lat, lng = _resolve_latlng(
                    cell(row, 'latitude'), cell(row, 'longitude'), _ref_lat, _ref_lng
                )
                if lat is None or lng is None:
                    skipped += 1
                    if len(skip_notes) < 10:
                        skip_notes.append(f'Row {i}: missing/invalid lat or lng')
                    continue
                if not (-90 <= lat <= 90) or not (-180 <= lng <= 180):
                    skipped += 1
                    if len(skip_notes) < 10:
                        skip_notes.append(f'Row {i}: out-of-range coordinates')
                    continue

                try:
                    acc_m = float((cell(row, 'accuracy_m') or '0').replace(',', '.'))
                except Exception:
                    acc_m = 0
                source = (cell(row, 'source') or 'import').strip().lower()[:20] or 'import'
                address = (cell(row, 'address') or '').strip()[:400]
                name = (cell(row, 'name') or '').strip() or 'Imported respondent'

                # Build the canonical geo JSON our parse_geo_answer understands
                geo_payload = json.dumps({
                    'lat': lat, 'lng': lng,
                    'accuracy_m': acc_m, 'source': source,
                    'address': address, 'captured_at': '',
                })

                response = SurveyResponse.objects.create(
                    survey=survey,
                    respondent=None,
                    respondent_name=name[:120],
                    is_public_response=False,
                )
                SurveyAnswer.objects.create(
                    response=response,
                    question=question,
                    value=geo_payload,
                    import_batch=batch,
                )
                created += 1

            batch.row_count = created
            batch.skipped_count = skipped
            batch.notes = '\n'.join(skip_notes)
            batch.save(update_fields=['row_count', 'skipped_count', 'notes'])

        return {
            'ok': True,
            'import_id': str(batch.pk),
            'created': created,
            'skipped': skipped,
            'skipped_notes': skip_notes,
            'redirect_url': reverse('survey_dashboard',
                                    kwargs={'pk': project.pk, 'sid': survey.pk}),
        }


# ─────────────────────────────────────────────────────────────────────────────
# 4. Pick-from-Files — list SharedFiles the user can choose as import source
# ─────────────────────────────────────────────────────────────────────────────

class ImportFilePickerView(LoginRequiredMixin, View):
    """Return a JSON list of CSV / XLSX files from the Files app."""

    ALLOWED_EXTS = {'csv', 'xlsx', 'xls'}

    def get(self, request, pk):
        from apps.files.models import SharedFile
        from django.db.models import Q

        user = request.user
        profile = getattr(user, 'staffprofile', None)
        unit = profile.unit if profile else None
        dept = profile.department if profile else None

        qs = SharedFile.objects.filter(
            Q(uploaded_by=user) |
            Q(visibility='office') |
            Q(visibility='unit', unit=unit) |
            Q(visibility='department', department=dept) |
            Q(shared_with=user) |
            Q(share_access__user=user)
        ).distinct().order_by('-created_at')

        files = []
        for f in qs:
            ext = os.path.splitext(f.name)[1].lower().lstrip('.')
            if ext not in self.ALLOWED_EXTS:
                continue
            files.append({
                'id': str(f.pk),
                'name': f.name,
                'ext': ext,
                'size': f.file_size,
                'size_display': f.size_display,
                'uploaded_at': f.created_at.strftime('%d %b %Y'),
                'project_code': str(f.project.code) if f.project else '',
            })

        return JsonResponse({'ok': True, 'files': files})


# ─────────────────────────────────────────────────────────────────────────────
# 5. Import management — list / unlink / bulk-delete / re-run
# ─────────────────────────────────────────────────────────────────────────────

class ProjectImportsView(LoginRequiredMixin, View):
    """GET — full management page listing every import on this project.

    Also accepts POST with action=unlink|delete|rerun&import_id=<uuid>.
    """

    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        imports = (
            LocationImport.objects
            .filter(project=project)
            .select_related('source_file', 'imported_by', 'survey', 'survey_question')
            .order_by('-created_at')
        )

        # Annotate live counts — how many rows are STILL in the DB (may be fewer
        # than batch.row_count if some were individually deleted)
        annotated = []
        for imp in imports:
            if imp.target == imp.Target.PROJECT_LOCATIONS:
                live = ProjectLocation.objects.filter(import_batch=imp).count()
            else:
                from .models import SurveyAnswer
                live = SurveyAnswer.objects.filter(import_batch=imp).count()
            annotated.append({
                'imp': imp,
                'live_count': live,
                'reduced': live < imp.row_count,
            })

        is_manager = request.user.is_superuser or project.project_manager == request.user
        return render(request, 'projects/location_imports.html', {
            'project': project,
            'imports': annotated,
            'is_manager': is_manager,
        })

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return JsonResponse({'error': 'Forbidden'}, status=403)
        is_manager = request.user.is_superuser or project.project_manager == request.user

        action = (request.POST.get('action') or '').strip()
        imp_id = request.POST.get('import_id') or ''
        try:
            imp = LocationImport.objects.get(pk=imp_id, project=project)
        except LocationImport.DoesNotExist:
            return JsonResponse({'error': 'Import not found'}, status=404)

        if action == 'unlink':
            # Clear the FK but keep all rows. History is preserved via
            # source_filename which was stored at import time.
            imp.source_file = None
            imp.save(update_fields=['source_file'])
            return JsonResponse({'ok': True, 'message': 'File unlinked. Data preserved.'})

        if action == 'delete':
            # Only the importer or a project manager may bulk-delete
            if not (is_manager or imp.imported_by_id == request.user.id):
                return JsonResponse({'error': 'Not permitted'}, status=403)
            with transaction.atomic():
                if imp.target == imp.Target.PROJECT_LOCATIONS:
                    ProjectLocation.objects.filter(import_batch=imp).delete()
                else:
                    from .models import SurveyAnswer, SurveyResponse
                    # Delete answers AND the responses that were created just for them
                    resp_ids = SurveyAnswer.objects.filter(
                        import_batch=imp).values_list('response_id', flat=True)
                    SurveyAnswer.objects.filter(import_batch=imp).delete()
                    # Only delete the response if it has no remaining answers
                    SurveyResponse.objects.filter(
                        pk__in=list(resp_ids)
                    ).filter(answers__isnull=True).delete()
                deleted_rows = imp.row_count
                imp.delete()
            return JsonResponse({
                'ok': True,
                'message': f'Deleted {deleted_rows} rows from this import.'
            })

        if action == 'rerun':
            # Re-run only possible if source_file is still linked
            if not imp.source_file_id:
                return JsonResponse({
                    'error': 'Source file is no longer linked; cannot re-run.'}, status=400)
            # We don't actually re-commit here — the client redirects to the
            # import modal with the file and saved mapping pre-filled.
            return JsonResponse({
                'ok': True,
                'shared_file_id': str(imp.source_file_id),
                'column_mapping': imp.column_mapping,
                'target': imp.target,
                'survey_id': str(imp.survey_id) if imp.survey_id else '',
                'question_id': str(imp.survey_question_id) if imp.survey_question_id else '',
            })

        return JsonResponse({'error': 'Unknown action'}, status=400)
# ---------------------------------------------------------------------------
# Project List
# ---------------------------------------------------------------------------

class ProjectListView(LoginRequiredMixin, TemplateView):
    template_name = 'projects/project_list.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        view = self.request.GET.get('view', 'my')
        q = self.request.GET.get('q', '')
        status = self.request.GET.get('status', '')
        priority = self.request.GET.get('priority', '')

        if user.is_superuser or view == 'all':
            projects = Project.objects.filter(
                Q(is_public=True) | Q(team_members=user) | Q(project_manager=user)
            ).distinct()
        else:
            projects = Project.objects.filter(
                Q(team_members=user) | Q(project_manager=user)
            ).distinct()

        if q:
            projects = projects.filter(
                Q(name__icontains=q) | Q(code__icontains=q) | Q(description__icontains=q)
            )
        if status:
            projects = projects.filter(status=status)
        if priority:
            projects = projects.filter(priority=priority)

        projects = projects.select_related('project_manager', 'department').order_by('-updated_at')

        my_projects = Project.objects.filter(
            Q(team_members=user) | Q(project_manager=user)
        ).distinct()

        ctx.update({
            'projects': projects,
            'status_choices': Project.Status.choices,
            'priority_choices': Project.Priority.choices,
            'view': view,
            'q': q,
            'status_filter': status,
            'priority_filter': priority,
            'total_count': projects.count(),
            'active_count': my_projects.filter(status='active').count(),
            'completed_count': my_projects.filter(status='completed').count(),
            'my_count': my_projects.count(),
        })
        return ctx


# ---------------------------------------------------------------------------
# Project Detail
# ---------------------------------------------------------------------------

class ProjectDetailView(LoginRequiredMixin, DetailView):
    model = Project
    template_name = 'projects/project_detail.html'
    context_object_name = 'project'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        project = self.object
        user = self.request.user

        milestones = project.milestones.all().order_by('order', 'due_date')
        tasks = project.tasks.select_related('assigned_to').order_by('status', '-created_at')[:15]
        updates = project.updates.select_related('author').order_by('-created_at')[:15]
        risks = project.risks.select_related('owner')
        team = project.team_members.filter(is_active=True).select_related(
            'staffprofile', 'staffprofile__position'
        )
        if project.project_manager_id:
            team = team.exclude(id=project.project_manager_id)

        total_ms = milestones.count()
        completed_ms = milestones.filter(status='completed').count()
        overdue_ms = milestones.filter(
            status__in=['pending', 'in_progress'],
            due_date__lt=timezone.now().date()
        ).count()

        budget_pct = 0
        if project.budget and project.budget > 0:
            budget_pct = min(100, int((project.budget_spent / project.budget) * 100))

        days_remaining = None
        days_remaining_abs = None
        is_overdue = False
        if project.end_date:
            delta = project.end_date - timezone.now().date()
            days_remaining = delta.days
            is_overdue = days_remaining < 0 and project.status not in ('completed', 'cancelled')
            days_remaining_abs = abs(days_remaining)

        project_tags = []
        if project.tags:
            project_tags = [tag.strip() for tag in project.tags.split(",") if tag.strip()]

        # ── Linked project channel ─────────────────────────────────────────────
        project_channel = ChatRoom.objects.filter(
            project=project,
            room_type='project',
            is_archived=False,
        ).first()

        # ── Linked project documents ───────────────────────────────────────
        try:
            from apps.files.models import SharedFile

            profile = getattr(user, 'staffprofile', None)
            unit = profile.unit if profile else None
            dept = profile.department if profile else None

            project_documents = SharedFile.objects.filter(
                project=project,
                is_latest=True,
            ).filter(
                Q(uploaded_by=user) |
                Q(visibility='office') |
                Q(visibility='unit', unit=unit) |
                Q(visibility='department', department=dept) |
                Q(shared_with=user) |
                Q(share_access__user=user) |
                Q(project__project_manager=user) |
                Q(project__team_members=user) |
                Q(project__is_public=True)
            ).distinct().select_related('uploaded_by', 'folder').order_by('-created_at')

            recent_project_documents = project_documents[:8]
            project_documents_count = project_documents.count()
        except Exception:
            project_documents = []
            recent_project_documents = []
            project_documents_count = 0

        # ── Linked meetings ─────────────────────────────────────────────
        try:
            from apps.meetings.models import Meeting
            project_meetings = Meeting.objects.filter(project=project).select_related(
                'organizer', 'project'
            ).order_by('-start_datetime')
            upcoming_meetings = project_meetings.filter(
                start_datetime__gte=timezone.now()
            ).order_by('start_datetime')[:5]
            recent_meetings = project_meetings[:5]
            meeting_count = project_meetings.count()
        except Exception:
            project_meetings = []
            upcoming_meetings = []
            recent_meetings = []
            meeting_count = 0

        # ── Linked finance data ─────────────────────────────────────────
        purchase_requests = project.purchase_requests.select_related(
            'requested_by', 'budget', 'approved_by'
        ).order_by('-created_at')[:8]

        try:
            from apps.finance.models import Payment
            project_payments = Payment.objects.filter(project=project).select_related(
                'paid_by', 'budget', 'purchase_request'
            ).order_by('-payment_date', '-created_at')[:8]
            finance_totals = Payment.objects.filter(project=project).aggregate(
                total_spent=Sum('amount')
            )
        except Exception:
            project_payments = []
            finance_totals = {'total_spent': 0}

        try:
            from apps.finance.models import EmployeeFinanceRequest
            employee_finance_requests = EmployeeFinanceRequest.objects.filter(
                project=project
            ).select_related('employee', 'budget', 'approved_by').order_by('-created_at')[:8]
        except Exception:
            employee_finance_requests = []

        total_purchase_est = purchase_requests.aggregate(t=Sum('estimated_cost'))['t'] if purchase_requests else 0
        finance_totals['total_purchase_est'] = total_purchase_est or 0

        ctx.update({
            'milestones': milestones,
            'updates': updates,
            'risks': risks,
            'active_risks': risks.filter(is_resolved=False),
            'resolved_risks': risks.filter(is_resolved=True),
            'tasks': tasks,
            'team_members': team,
            'is_member': _is_project_member(user, project),
            'is_manager': user.is_superuser or project.project_manager == user,
            'can_manage_milestones': _is_project_owner(user, project),
            'can_post_update': _is_project_owner(user, project),

            'total_ms': total_ms,
            'completed_ms': completed_ms,
            'overdue_ms': overdue_ms,
            'budget_pct': budget_pct,
            'days_remaining': days_remaining,
            'days_remaining_abs': days_remaining_abs,
            'is_overdue': is_overdue,
            'project_tags': project_tags,
            'project_channel': project_channel,

            'status_choices': Project.Status.choices,
            'risk_levels': Risk.Level.choices,
            'milestone_statuses': Milestone.Status.choices,
            'staff_list': User.objects.filter(status='active', is_active=True).order_by('first_name'),

            # finance
            'purchase_requests': purchase_requests,
            'project_payments': project_payments,
            'employee_finance_requests': employee_finance_requests,
            'finance_totals': finance_totals,

            # meetings
            'project_meetings': project_meetings,
            'upcoming_meetings': upcoming_meetings,
            'recent_meetings': recent_meetings,
            'meeting_count': meeting_count,

            # Document
            'project_documents': project_documents,
            'recent_project_documents': recent_project_documents,
            'project_documents_count': project_documents_count,
        })
        return ctx


# ---------------------------------------------------------------------------
# Project Create
# ---------------------------------------------------------------------------

class ProjectCreateView(LoginRequiredMixin, View):
    template_name = 'projects/project_form.html'

    def get(self, request):
        from apps.organization.models import Department, Unit
        return render(request, self.template_name, {
            'departments': Department.objects.filter(is_active=True),
            'units': Unit.objects.filter(is_active=True),
            'staff_list': User.objects.filter(status='active', is_active=True).order_by('first_name'),
            'priority_choices': Project.Priority.choices,
            'status_choices': Project.Status.choices,
            'mode': 'create',
        })

    def post(self, request):
        try:
            import shortuuid
            code = request.POST.get('code', '').strip() or shortuuid.ShortUUID().random(length=6).upper()
        except ImportError:
            import random, string
            code = request.POST.get('code', '').strip() or ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

        project = Project.objects.create(
            name=request.POST['name'],
            code=code,
            description=request.POST.get('description', ''),
            project_manager_id=request.POST.get('project_manager') or request.user.id,
            priority=request.POST.get('priority', 'medium'),
            status=request.POST.get('status', 'planning'),
            start_date=request.POST.get('start_date') or None,
            end_date=request.POST.get('end_date') or None,
            budget=request.POST.get('budget') or None,
            cover_color=request.POST.get('cover_color', '#1e3a5f'),
            tags=request.POST.get('tags', ''),
            is_public=request.POST.get('is_public') == 'on',
        )
        member_ids = request.POST.getlist('team_members')
        if member_ids:
            project.team_members.set(User.objects.filter(id__in=member_ids))
        if request.POST.get('dept_id'):
            try:
                from apps.organization.models import Department
                project.department = Department.objects.get(id=request.POST['dept_id'])
                project.save(update_fields=['department'])
            except Exception:
                pass

        messages.success(request, f'Project "{project.name}" created.')
        return redirect('project_detail', pk=project.id)


# ---------------------------------------------------------------------------
# Project Edit
# ---------------------------------------------------------------------------

class ProjectEditView(LoginRequiredMixin, View):
    template_name = 'projects/project_form.html'

    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            messages.error(request, 'You do not have permission to edit this project.')
            return redirect('project_detail', pk=pk)
        from apps.organization.models import Department, Unit
        return render(request, self.template_name, {
            'project': project,
            'departments': Department.objects.filter(is_active=True),
            'units': Unit.objects.filter(is_active=True),
            'staff_list': User.objects.filter(status='active', is_active=True).order_by('first_name'),
            'priority_choices': Project.Priority.choices,
            'status_choices': Project.Status.choices,
            'mode': 'edit',
        })

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            messages.error(request, 'Permission denied.')
            return redirect('project_detail', pk=pk)

        project.name = request.POST.get('name', project.name)
        project.description = request.POST.get('description', project.description)
        project.priority = request.POST.get('priority', project.priority)
        project.status = request.POST.get('status', project.status)
        project.start_date = request.POST.get('start_date') or None
        project.end_date = request.POST.get('end_date') or None
        project.budget = request.POST.get('budget') or None
        project.cover_color = request.POST.get('cover_color', project.cover_color)
        project.tags = request.POST.get('tags', project.tags)
        project.is_public = request.POST.get('is_public') == 'on'
        pm_id = request.POST.get('project_manager')
        if pm_id:
            try:
                project.project_manager = User.objects.get(id=pm_id)
            except User.DoesNotExist:
                pass
        if request.POST.get('dept_id'):
            try:
                from apps.organization.models import Department
                project.department = Department.objects.get(id=request.POST['dept_id'])
            except Exception:
                pass
        project.save()

        member_ids = request.POST.getlist('team_members')
        if member_ids is not None:
            project.team_members.set(User.objects.filter(id__in=member_ids))

        messages.success(request, f'Project "{project.name}" updated.')
        return redirect('project_detail', pk=pk)


# ---------------------------------------------------------------------------
# Post Update / Change Status + Progress
# ---------------------------------------------------------------------------

class ProjectUpdateStatusView(LoginRequiredMixin, View):
    """
    Post a status/progress update. Only the project manager and superusers
    can post here — keeping ownership of the project narrative in their
    hands. Other members can still comment via the chat / discussion area.

    Optionally accepts `milestone` (UUID of a milestone in this project)
    so manual updates can be tied to a milestone and surface in its
    history alongside the auto-generated completion notes.
    """

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)

        # Permission check tightened: only PM / superuser. Was previously
        # _is_project_member which let any team member move the project
        # status — that wasn't the intent.
        if not _is_project_owner(request.user, project):
            messages.error(
                request,
                'Only the project manager or office admin can update the '
                'project status.'
            )
            return redirect('project_detail', pk=pk)

        content = request.POST.get('content', '').strip()
        new_status = request.POST.get('status', project.status)
        try:
            progress = max(0, min(100, int(request.POST.get('progress_pct', project.progress_pct))))
        except (ValueError, TypeError):
            progress = project.progress_pct

        # Optional: tie this update to a milestone in the same project.
        milestone_obj = None
        milestone_id = (request.POST.get('milestone') or '').strip()
        if milestone_id:
            milestone_obj = (
                Milestone.objects
                .filter(pk=milestone_id, project=project)
                .first()
            )
            if milestone_obj is None:
                messages.error(request, 'That milestone does not belong to this project.')
                return redirect('project_detail', pk=pk)

        if content:
            ProjectUpdate.objects.create(
                project=project,
                author=request.user,
                title=request.POST.get('title', ''),
                content=content,
                progress_at_time=progress,
                status_at_time=new_status,
                milestone=milestone_obj,
                is_milestone_update=bool(milestone_obj),
            )

        project.status = new_status
        project.progress_pct = progress
        if new_status == 'completed' and not project.completed_at:
            project.completed_at = timezone.now()
        project.save(update_fields=[
            'status', 'progress_pct', 'completed_at', 'updated_at',
        ])
        messages.success(request, 'Project updated.')
        return redirect('project_detail', pk=pk)

# ---------------------------------------------------------------------------
# Add Milestone
# ---------------------------------------------------------------------------

class ProjectMilestoneView(LoginRequiredMixin, View):
    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)

        if not _is_project_member(request.user, project):
            messages.error(request, 'Permission denied.')
            return redirect('project_detail', pk=pk)

        assigned_to_id = request.POST.get('assigned_to') or None

        if assigned_to_id and not _is_project_owner(request.user, project):
            messages.error(request, 'Only the project owner/manager can assign milestone owners.')
            return redirect('project_detail', pk=pk)

        milestone = Milestone.objects.create(
            project=project,
            name=request.POST['name'],
            description=request.POST.get('description', ''),
            due_date=request.POST['due_date'],
            assigned_to_id=assigned_to_id,
            created_by=request.user,
            order=project.milestones.count() + 1,
        )

        if assigned_to_id:
            project.team_members.add(milestone.assigned_to)
            _create_or_update_milestone_task(milestone, request.user)
            messages.success(request, 'Milestone added and linked task created.')
        else:
            messages.success(request, 'Milestone added.')

        return redirect('project_detail', pk=pk)


# ---------------------------------------------------------------------------
# Update Milestone Status
# ---------------------------------------------------------------------------

class MilestoneStatusView(LoginRequiredMixin, View):
    def post(self, request, pk, mid):
        project = get_object_or_404(Project, pk=pk)
        milestone = get_object_or_404(Milestone, pk=mid, project=project)

        if not _is_project_member(request.user, project):
            messages.error(request, 'Permission denied.')
            return redirect('project_detail', pk=pk)

        is_owner = _is_project_owner(request.user, project)
        changed_fields = []

        new_status = request.POST.get('status')
        if new_status in dict(Milestone.Status.choices):
            milestone.status = new_status
            changed_fields.append('status')

            if new_status == Milestone.Status.COMPLETED and not milestone.completed_at:
                milestone.completed_at = timezone.now()
                changed_fields.append('completed_at')
            elif new_status != Milestone.Status.COMPLETED and milestone.completed_at:
                milestone.completed_at = None
                changed_fields.append('completed_at')

        if 'due_date' in request.POST:
            if not is_owner:
                messages.error(request, 'Only the project owner/manager can adjust milestone dates.')
                return redirect('project_detail', pk=pk)

            due_date = request.POST.get('due_date')
            if due_date:
                milestone.due_date = due_date
                changed_fields.append('due_date')

        if 'assigned_to' in request.POST:
            if not is_owner:
                messages.error(request, 'Only the project owner/manager can assign milestone owners.')
                return redirect('project_detail', pk=pk)

            assigned_to_id = request.POST.get('assigned_to') or None
            milestone.assigned_to_id = assigned_to_id
            changed_fields.append('assigned_to')

        if changed_fields:
            if 'updated_at' not in changed_fields:
                changed_fields.append('updated_at')

            milestone.save(update_fields=list(set(changed_fields)))

            if milestone.assigned_to_id:
                project.team_members.add(milestone.assigned_to)
                _create_or_update_milestone_task(milestone, request.user)

        messages.success(request, f'Milestone "{milestone.name}" updated.')
        return redirect('project_detail', pk=pk)


# ---------------------------------------------------------------------------
# Reorder Milestones (drag-and-drop)
# ---------------------------------------------------------------------------

class MilestoneReorderView(LoginRequiredMixin, View):
    """
    Accepts a JSON POST: {"order": ["<milestone_uuid>", "<milestone_uuid>", ...]}
    Reorders milestones for the project. Disabled once the project is closed
    (status = completed or cancelled).
    """

    CLOSED_STATUSES = {'completed', 'cancelled'}

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)

        if not _is_project_member(request.user, project):
            return JsonResponse({'ok': False, 'error': 'Permission denied.'}, status=403)

        if project.status in self.CLOSED_STATUSES:
            return JsonResponse(
                {'ok': False, 'error': 'Project is closed — milestones cannot be reordered.'},
                status=400,
            )

        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
            new_order = payload.get('order') or []
            if not isinstance(new_order, list):
                raise ValueError('order must be a list')
        except (ValueError, json.JSONDecodeError):
            return JsonResponse({'ok': False, 'error': 'Invalid payload.'}, status=400)

        # Only touch milestones that belong to this project
        valid_ids = set(
            str(mid) for mid in project.milestones.values_list('id', flat=True)
        )
        clean_order = [mid for mid in new_order if str(mid) in valid_ids]

        if not clean_order:
            return JsonResponse({'ok': False, 'error': 'No valid milestones provided.'}, status=400)

        with transaction.atomic():
            for index, mid in enumerate(clean_order, start=1):
                Milestone.objects.filter(pk=mid, project=project).update(order=index)

        return JsonResponse({'ok': True, 'count': len(clean_order)})

# ---------------------------------------------------------------------------
# Risks
# ---------------------------------------------------------------------------

class ProjectRiskView(LoginRequiredMixin, View):
    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            messages.error(request, 'Permission denied.')
            return redirect('project_detail', pk=pk)

        action = request.POST.get('action', 'add')

        if action == 'resolve':
            risk = get_object_or_404(Risk, pk=request.POST.get('risk_id'), project=project)
            risk.is_resolved = True
            risk.save(update_fields=['is_resolved'])
            messages.success(request, f'Risk "{risk.title}" marked as resolved.')

        elif action == 'unresolve':
            risk = get_object_or_404(Risk, pk=request.POST.get('risk_id'), project=project)
            risk.is_resolved = False
            risk.save(update_fields=['is_resolved'])
            messages.success(request, f'Risk "{risk.title}" re-opened.')

        elif action == 'add':
            owner_id = request.POST.get('owner_id')
            Risk.objects.create(
                project=project,
                title=request.POST['title'],
                description=request.POST.get('description', ''),
                level=request.POST.get('level', 'medium'),
                mitigation=request.POST.get('mitigation', ''),
                owner_id=owner_id or None,
            )
            messages.success(request, 'Risk logged.')

        return redirect('project_detail', pk=pk)


class ProjectGanttPDFView(LoginRequiredMixin, View):
    """
    Generate and download a Gantt chart PDF for a project.
    The chart shows the project span and all milestones on a landscape A4 page.
    Requires: pip install reportlab
    """

    def get(self, request, pk):
        from apps.projects.models import Project

        project = get_object_or_404(Project, pk=pk)

        # Access control: only project members / public projects
        is_member = (
                project.is_public
                or project.project_manager == request.user
                or project.team_members.filter(pk=request.user.pk).exists()
                or request.user.is_superuser
        )
        if not is_member:
            return HttpResponseForbidden("You do not have access to this project.")

        milestones = list(project.milestones.all().order_by('order', 'due_date'))
        pdf_bytes = _build_gantt_pdf(project, milestones)

        filename = f"{project.code}-gantt.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ─────────────────────────────────────────────────────────────────────────────
# _build_gantt_pdf helper
# ─────────────────────────────────────────────────────────────────────────────

def _build_gantt_pdf(project, milestones):
    """
    Build a landscape A4 Gantt chart PDF using reportlab.
    Returns raw PDF bytes.
    """
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
    except ImportError:
        raise RuntimeError("reportlab is not installed. Run: pip install reportlab")

    buf = io.BytesIO()

    PAGE_W, PAGE_H = landscape(A4)  # ~842 × 595 pt
    MARGIN = 18 * mm

    # ── Layout constants ──────────────────────────────────────────────────────
    LABEL_W = 148  # left column (milestone names)
    CHART_X = MARGIN + LABEL_W + 6  # x start of the timeline area
    CHART_W = PAGE_W - CHART_X - MARGIN  # timeline width
    CHART_TOP = PAGE_H - MARGIN - 50  # y of the top edge (reportlab: bottom-left origin)

    TITLE_H = 28
    META_H = 16
    HDR_H = 22  # month strip height
    PROJ_H = 16  # project bar row height
    ROW_H = 24  # each milestone row height
    LEGEND_H = 14

    today = date.today()

    # ── Determine timeline range ───────────────────────────────────────────────
    ms_dates = [m.due_date for m in milestones if m.due_date]

    t_min = project.start_date or (min(ms_dates) if ms_dates else today - timedelta(days=30))
    t_max = project.end_date or (max(ms_dates) if ms_dates else today + timedelta(days=90))

    # Extend range to include today + 1-week padding on each side
    t_min = date(t_min.year, t_min.month, 1)
    t_max_padded = t_max + timedelta(days=14)
    t_max_month_end = date(
        t_max_padded.year,
        t_max_padded.month,
        1
    )
    # Go to start of following month for a clean right edge
    if t_max_month_end.month == 12:
        t_max = date(t_max_month_end.year + 1, 1, 1)
    else:
        t_max = date(t_max_month_end.year, t_max_month_end.month + 1, 1)

    total_days = max(1, (t_max - t_min).days)

    def x_of(d):
        """X coordinate for a date."""
        return CHART_X + (d - t_min).days / total_days * CHART_W

    # ── Colours ───────────────────────────────────────────────────────────────
    C_COVER = colors.HexColor(project.cover_color or '#1e3a5f')
    C_PENDING = colors.HexColor('#94a3b8')
    C_IN_PROG = colors.HexColor('#3b82f6')
    C_DONE = colors.HexColor('#10b981')
    C_MISSED = colors.HexColor('#ef4444')
    C_TODAY = colors.HexColor('#f59e0b')
    C_GRID = colors.HexColor('#e2e8f0')
    C_BG_ALT = colors.HexColor('#f8fafc')
    C_TEXT = colors.HexColor('#1e293b')
    C_MUTED = colors.HexColor('#64748b')
    C_WHITE = colors.white
    C_LABEL_COL = colors.HexColor('#f1f5f9')  # left column background

    STATUS_COLOR = {
        'pending': C_PENDING,
        'in_progress': C_IN_PROG,
        'completed': C_DONE,
        'missed': C_MISSED,
    }

    c = rl_canvas.Canvas(buf, pagesize=landscape(A4))
    c.setTitle(f"{project.name} — Gantt Chart")

    # ── Total chart height calc ───────────────────────────────────────────────
    n_rows = len(milestones)
    chart_h = HDR_H + PROJ_H + n_rows * ROW_H  # just the rows

    # y_top: top of the Gantt rows block (not title)
    title_block = TITLE_H + META_H + 6
    y_top = PAGE_H - MARGIN - title_block - 4  # top of month header

    # ── Title block ──────────────────────────────────────────────────────────
    # Colour bar at very top
    c.setFillColor(C_COVER)
    c.rect(MARGIN, PAGE_H - MARGIN - TITLE_H, PAGE_W - 2 * MARGIN, TITLE_H, stroke=0, fill=1)

    c.setFillColor(C_WHITE)
    c.setFont('Helvetica-Bold', 13)
    c.drawString(MARGIN + 8, PAGE_H - MARGIN - TITLE_H + 8, f"{project.name}  —  Gantt Chart")

    c.setFont('Helvetica', 8)
    c.setFillColor(colors.HexColor('#dbeafe'))
    right_title = f"{project.code}  |  {project.get_status_display()}  |  Progress: {project.progress_pct}%"
    c.drawRightString(PAGE_W - MARGIN - 8, PAGE_H - MARGIN - TITLE_H + 8, right_title)

    # Meta line
    c.setFillColor(C_MUTED)
    c.setFont('Helvetica', 8)
    meta_y = PAGE_H - MARGIN - TITLE_H - META_H + 2
    start_str = project.start_date.strftime('%d %b %Y') if project.start_date else 'TBD'
    end_str = project.end_date.strftime('%d %b %Y') if project.end_date else 'TBD'
    meta_txt = f"Start: {start_str}   End: {end_str}   Milestones: {n_rows}   Generated: {today.strftime('%d %b %Y')}"
    c.drawString(MARGIN, meta_y, meta_txt)

    # Separator
    c.setStrokeColor(C_GRID)
    c.setLineWidth(0.5)
    c.line(MARGIN, meta_y - 3, PAGE_W - MARGIN, meta_y - 3)

    # ── Month header row ──────────────────────────────────────────────────────
    hdr_y = y_top  # bottom of header row
    hdr_top = y_top + HDR_H  # top of header row

    c.setFillColor(C_LABEL_COL)
    c.rect(MARGIN, hdr_y, LABEL_W + 6, HDR_H, stroke=0, fill=1)
    c.setFillColor(C_MUTED)
    c.setFont('Helvetica-Bold', 7)
    c.drawString(MARGIN + 4, hdr_y + HDR_H / 2 - 3, "MILESTONE")

    # Month columns
    cur = date(t_min.year, t_min.month, 1)
    while cur < t_max:
        mx = x_of(cur)
        # Vertical grid line
        c.setStrokeColor(C_GRID)
        c.setLineWidth(0.4)
        c.line(mx, MARGIN, mx, hdr_top)

        # Month label in header
        c.setFillColor(C_MUTED)
        c.setFont('Helvetica-Bold', 7)
        label = cur.strftime('%b %Y')
        c.drawString(mx + 3, hdr_y + 7, label)

        # Advance to next month
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)

    # Header bottom border
    c.setStrokeColor(C_MUTED)
    c.setLineWidth(0.8)
    c.line(MARGIN, hdr_y, PAGE_W - MARGIN, hdr_y)

    # ── Project span row ──────────────────────────────────────────────────────
    proj_y = hdr_y - PROJ_H  # bottom of this row

    # Row background
    c.setFillColor(colors.HexColor('#f0f4ff'))
    c.rect(MARGIN, proj_y, PAGE_W - 2 * MARGIN, PROJ_H, stroke=0, fill=1)

    # Row label
    c.setFillColor(C_TEXT)
    c.setFont('Helvetica-Bold', 7.5)
    c.drawString(MARGIN + 4, proj_y + PROJ_H / 2 - 3, "Project Span")

    # Project bar
    if project.start_date and project.end_date:
        ps_x = x_of(project.start_date)
        pe_x = x_of(project.end_date)
        bar_w = max(2, pe_x - ps_x)
        bar_y = proj_y + 4
        bar_h = PROJ_H - 8

        # Track
        c.setFillColor(colors.HexColor('#dbeafe'))
        c.roundRect(ps_x, bar_y, bar_w, bar_h, 3, stroke=0, fill=1)

        # Progress fill
        prog_w = max(0, bar_w * project.progress_pct / 100)
        if prog_w > 0:
            c.setFillColor(C_COVER)
            c.roundRect(ps_x, bar_y, prog_w, bar_h, 3, stroke=0, fill=1)

        # Progress label
        c.setFillColor(C_COVER)
        c.setFont('Helvetica-Bold', 7)
        c.drawString(ps_x + 4, bar_y + bar_h / 2 - 3, f"{project.progress_pct}%")

    # Row border
    c.setStrokeColor(C_GRID)
    c.setLineWidth(0.5)
    c.line(MARGIN, proj_y, PAGE_W - MARGIN, proj_y)

    # ── Milestone rows ────────────────────────────────────────────────────────
    for idx, ms in enumerate(milestones):
        row_y = proj_y - (idx + 1) * ROW_H  # bottom of this row
        sc = STATUS_COLOR.get(ms.status, C_PENDING)

        # Alternating background
        if idx % 2 == 1:
            c.setFillColor(C_BG_ALT)
            c.rect(MARGIN, row_y, PAGE_W - 2 * MARGIN, ROW_H, stroke=0, fill=1)

        # ── Label column ─────────────────────────────────────────────────────
        # Number badge
        badge_x = MARGIN + 4
        badge_y = row_y + (ROW_H - 10) / 2
        c.setFillColor(sc)
        c.roundRect(badge_x, badge_y, 12, 10, 2, stroke=0, fill=1)
        c.setFillColor(C_WHITE)
        c.setFont('Helvetica-Bold', 6)
        c.drawCentredString(badge_x + 6, badge_y + 2.5, str(idx + 1))

        # Milestone name (truncate to fit)
        max_chars = 22
        name_disp = ms.name if len(ms.name) <= max_chars else ms.name[:max_chars] + '…'
        c.setFillColor(C_TEXT)
        c.setFont('Helvetica-Bold', 7.5)
        c.drawString(badge_x + 16, row_y + ROW_H / 2 + 1, name_disp)

        # Status text
        status_labels = {
            'pending': 'Pending', 'in_progress': 'In Progress',
            'completed': 'Completed', 'missed': 'Missed'
        }
        c.setFillColor(sc)
        c.setFont('Helvetica', 6.5)
        c.drawString(badge_x + 16, row_y + ROW_H / 2 - 7, status_labels.get(ms.status, ''))

        # ── Timeline area ─────────────────────────────────────────────────────
        if not ms.due_date:
            c.setStrokeColor(C_GRID)
            c.setLineWidth(0.3)
            c.line(MARGIN, row_y, PAGE_W - MARGIN, row_y)
            continue

        due_x = x_of(ms.due_date)
        bar_y = row_y + (ROW_H - 6) / 2
        bar_h = 6
        track_x = CHART_X + 2

        # Full grey track
        c.setFillColor(C_GRID)
        c.roundRect(track_x, bar_y, CHART_W - 4, bar_h, 3, stroke=0, fill=1)

        # Coloured fill up to due date
        fill_w = max(0, due_x - track_x - 8)
        if fill_w > 0:
            sc_light = colors.HexColor(
                {'pending': '#e2e8f0', 'in_progress': '#dbeafe',
                 'completed': '#d1fae5', 'missed': '#fee2e2'}.get(ms.status, '#e2e8f0')
            )
            c.setFillColor(sc_light)
            c.roundRect(track_x, bar_y, fill_w, bar_h, 3, stroke=0, fill=1)

        # Diamond marker at due date
        d_cx = due_x
        d_cy = row_y + ROW_H / 2
        d_sz = 5.5
        from reportlab.graphics.shapes import Polygon
        # Draw diamond using lines
        c.setFillColor(sc)
        c.setStrokeColor(C_WHITE)
        c.setLineWidth(1)
        path = c.beginPath()
        path.moveTo(d_cx, d_cy + d_sz)
        path.lineTo(d_cx + d_sz, d_cy)
        path.lineTo(d_cx, d_cy - d_sz)
        path.lineTo(d_cx - d_sz, d_cy)
        path.close()
        c.drawPath(path, fill=1, stroke=1)

        # Status icon inside diamond
        icon = {'completed': '✓', 'missed': '✕', 'in_progress': '▶'}.get(ms.status, '·')
        c.setFillColor(C_WHITE)
        c.setFont('Helvetica-Bold', 5)
        c.drawCentredString(d_cx, d_cy - 1.8, icon)

        # Due-date label above diamond
        due_label = ms.due_date.strftime('%d %b')
        c.setFillColor(sc)
        c.setFont('Helvetica-Bold', 6.5)
        c.drawCentredString(d_cx, d_cy + d_sz + 3, due_label)

        # Row separator
        c.setStrokeColor(C_GRID)
        c.setLineWidth(0.3)
        c.line(MARGIN, row_y, PAGE_W - MARGIN, row_y)

    # ── Today line ────────────────────────────────────────────────────────────
    today_x = x_of(today)
    chart_bottom = proj_y - n_rows * ROW_H
    if CHART_X <= today_x <= CHART_X + CHART_W:
        c.setStrokeColor(C_TODAY)
        c.setLineWidth(1.2)
        c.setDash(4, 3)
        c.line(today_x, chart_bottom, today_x, hdr_top)
        c.setDash()  # reset dash

        # "TODAY" badge at top of line
        badge_w, badge_h = 28, 10
        c.setFillColor(C_TODAY)
        c.roundRect(today_x - badge_w / 2, hdr_top, badge_w, badge_h, 3, stroke=0, fill=1)
        c.setFillColor(C_WHITE)
        c.setFont('Helvetica-Bold', 6)
        c.drawCentredString(today_x, hdr_top + 3, 'TODAY')

    # ── Column separator ──────────────────────────────────────────────────────
    sep_x = CHART_X - 2
    c.setStrokeColor(colors.HexColor('#cbd5e1'))
    c.setLineWidth(1)
    c.line(sep_x, chart_bottom, sep_x, hdr_top)

    # Outer border
    c.setStrokeColor(C_GRID)
    c.setLineWidth(0.8)
    c.rect(MARGIN, chart_bottom, PAGE_W - 2 * MARGIN, hdr_top - chart_bottom, stroke=1, fill=0)

    # ── Legend strip at bottom of page ────────────────────────────────────────
    legend_y = MARGIN + 2
    c.setFont('Helvetica-Bold', 6.5)
    c.setFillColor(C_MUTED)
    c.drawString(MARGIN, legend_y, 'Legend:')
    lx = MARGIN + 36
    for lbl, col in [('Pending', C_PENDING), ('In Progress', C_IN_PROG),
                     ('Completed', C_DONE), ('Missed', C_MISSED),
                     ('Today', C_TODAY)]:
        c.setFillColor(col)
        c.circle(lx + 4, legend_y + 3, 3.5, fill=1, stroke=0)
        c.setFillColor(C_MUTED)
        c.setFont('Helvetica', 6.5)
        c.drawString(lx + 10, legend_y, lbl)
        lx += c.stringWidth(lbl, 'Helvetica', 6.5) + 22

    c.save()
    return buf.getvalue()


# =============================================================================
# Document Upload VIEWS
# =============================================================================

class ProjectDocumentLinkView(LoginRequiredMixin, View):
    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        q = request.GET.get('q', '').strip()

        profile = getattr(request.user, 'staffprofile', None)
        unit = profile.unit if profile else None
        dept = profile.department if profile else None

        files = SharedFile.objects.filter(
            Q(uploaded_by=request.user) |
            Q(visibility='office') |
            Q(visibility='unit', unit=unit) |
            Q(visibility='department', department=dept) |
            Q(shared_with=request.user) |
            Q(share_access__user=request.user)
        ).distinct().select_related('uploaded_by', 'folder').order_by('-created_at')

        if q:
            files = files.filter(
                Q(name__icontains=q) |
                Q(description__icontains=q) |
                Q(tags__icontains=q)
            )

        results = []
        for f in files[:30]:
            results.append({
                'id': str(f.id),
                'name': f.name,
                'description': f.description or '',
                'uploaded_by': f.uploaded_by.full_name if f.uploaded_by else '',
                'folder': f.folder.name if f.folder else '',
                'created_at': timezone.localtime(f.created_at).strftime('%d %b %Y %H:%M'),
                'size': getattr(f, 'size_display', ''),
                'icon_class': f.icon_class,
                'icon_color': f.icon_color,
                'is_linked': str(getattr(f, 'project_id', '') or '') == str(project.id),
            })

        return JsonResponse({'results': results})

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        file_id = request.POST.get('file_id', '').strip()
        if not file_id:
            messages.error(request, 'No file selected.')
            return redirect('project_detail', pk=pk)

        profile = getattr(request.user, 'staffprofile', None)
        unit = profile.unit if profile else None
        dept = profile.department if profile else None

        file_obj = get_object_or_404(
            SharedFile.objects.filter(
                Q(uploaded_by=request.user) |
                Q(visibility='office') |
                Q(visibility='unit', unit=unit) |
                Q(visibility='department', department=dept) |
                Q(shared_with=request.user) |
                Q(share_access__user=request.user)
            ).distinct(),
            pk=file_id
        )

        file_obj.project = project
        file_obj.save(update_fields=['project'])

        messages.success(request, f'"{file_obj.name}" linked to project.')
        return redirect('project_detail', pk=pk)


class ProjectDocumentUnlinkView(LoginRequiredMixin, View):
    def post(self, request, pk, file_id):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        file_obj = get_object_or_404(SharedFile, pk=file_id, project=project)
        file_obj.project = None
        file_obj.save(update_fields=['project'])

        messages.success(request, f'"{file_obj.name}" removed from this project.')
        return redirect('project_detail', pk=pk)


# =============================================================================
# SURVEY VIEWS
# =============================================================================

class ProjectSurveyListView(LoginRequiredMixin, ListView):
    model = Survey
    template_name = 'projects/project_surveys.html'
    context_object_name = 'surveys'

    def get_queryset(self):
        self.project = get_object_or_404(Project, pk=self.kwargs['pk'])
        return Survey.objects.filter(project=self.project).order_by('-created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['project'] = self.project
        ctx['is_member'] = _is_project_member(self.request.user, self.project)
        ctx['is_manager'] = (
            self.request.user.is_superuser or
            self.project.project_manager == self.request.user
        )
        # annotate each survey with a resp_count attribute the template uses
        for survey in ctx['surveys']:
            if not hasattr(survey, 'resp_count'):
                survey.resp_count = survey.responses.count()
        return ctx

    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        project = self.project

        export_sid = request.GET.get('export')
        if export_sid:
            return self._build_survey_export(request, project, export_sid)

        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, pk):
        """Create a new survey with its questions."""
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        title = request.POST.get('title', '').strip()
        if not title:
            messages.error(request, 'Survey title is required.')
            return redirect('project_surveys', pk=pk)

        import uuid as _uuid
        is_public = request.POST.get('is_public') == 'on'
        survey = Survey.objects.create(
            project=project,
            title=title,
            description=request.POST.get('description', '').strip(),
            is_anonymous=request.POST.get('is_anonymous') == 'on',
            is_public=is_public,
            allow_multiple_responses=request.POST.get('allow_multiple_responses') == 'on',
            closes_at=request.POST.get('closes_at') or None,
            created_by=request.user,
            public_token=_uuid.uuid4() if is_public else None,
        )

        _save_survey_questions(survey, request)

        messages.success(request, f'Survey "{survey.title}" created.')
        return redirect('survey_dashboard', pk=pk, sid=survey.pk)

    def _build_survey_export(self, request, project, export_sid):
        survey = get_object_or_404(
            Survey,
            pk=export_sid,
            project=project
        )

        questions = survey.questions.all().order_by('order')
        responses = survey.responses.all().order_by('-submitted_at')

        filename = f"{survey.title or 'survey'}_export.csv".replace(' ', '_')

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)

        header = [
            'Response ID',
            'Respondent Name',
            'Respondent',
            'Is Public',
            'Submitted At',
            'IP Address',
            'Device ID',
            'User Agent',
        ]

        for q in questions:
            header.append(q.text)

        writer.writerow(header)

        for resp in responses:
            row = [
                str(resp.pk),
                getattr(resp, 'respondent_name', '') or '',
                str(getattr(resp, 'respondent', '') or ''),
                getattr(resp, 'is_public_response', False),
                getattr(resp, 'submitted_at', ''),
                getattr(resp, 'ip_address', ''),
                getattr(resp, 'device_id', ''),
                getattr(resp, 'user_agent', ''),
            ]

            answers_by_question = {
                str(ans.question_id): ans
                for ans in resp.answers.select_related('question').all()
            }

            for q in questions:
                ans = answers_by_question.get(str(q.pk))
                if not ans:
                    row.append('')
                    continue

                value = getattr(ans, 'value', '') or ''

                if q.q_type == 'geolocation':
                    geo = parse_geo_answer(value)
                    if geo:
                        parts = []
                        if geo.get('address'):
                            parts.append(geo['address'])
                        if geo.get('lat') is not None and geo.get('lng') is not None:
                            parts.append(f"{geo['lat']:.6f}, {geo['lng']:.6f}")
                        if geo.get('accuracy_m'):
                            parts.append(f"±{float(geo['accuracy_m']):.0f}m")
                        if geo.get('source'):
                            parts.append(str(geo['source']).upper())
                        value = ' | '.join(parts) if parts else value

                row.append(value)

            writer.writerow(row)

        return response

class SurveyDetailView(LoginRequiredMixin, View):
    """Results dashboard. Redirects non-members to submit page."""

    def get(self, request, pk, sid):
        project = get_object_or_404(Project, pk=pk)
        survey = get_object_or_404(Survey, pk=sid, project=project)
        is_member = _is_project_member(request.user, project)
        is_manager = request.user.is_superuser or project.project_manager == request.user

        if not is_member:
            return redirect('survey_submit', pk=pk, sid=sid)

        questions = survey.questions.order_by('order')
        responses = (survey.responses
                     .select_related('respondent')
                     .prefetch_related('answers__question')
                     .order_by('-submitted_at'))

        already_responded = (
                not survey.is_anonymous
                and responses.filter(respondent=request.user).exists()
        )

        # ── Pre-process per-question results so templates need no custom filters ──
        question_results = []
        for q in questions:
            raw_answers = list(
                SurveyAnswer.objects.filter(question=q).values_list('value', flat=True)
            )
            result = {'question': q, 'raw': raw_answers, 'count': len(raw_answers)}

            if q.q_type in ('single_choice', 'multi_choice'):
                all_vals = []
                for a in raw_answers:
                    all_vals.extend([v.strip() for v in a.split(',') if v.strip()])
                tally = dict(Counter(all_vals))
                total = sum(tally.values()) or 1
                options = q.get_options_list()
                # Build ordered list safe for templates: [{label, count, pct}, ...]
                tally_list = []
                for opt in options:
                    cnt = tally.get(opt, 0)
                    tally_list.append({
                        'label': opt, 'count': cnt,
                        'pct': round(cnt / total * 100),
                    })
                # Any write-in answers not in the defined options
                for k, v in tally.items():
                    if k not in options:
                        tally_list.append({
                            'label': k, 'count': v,
                            'pct': round(v / total * 100), 'extra': True,
                        })
                result.update({
                    'tally_list': tally_list,
                    'total': sum(tally.values()),
                })

            elif q.q_type == 'rating':
                nums = [int(a) for a in raw_answers if a.strip().isdigit()]
                total = len(nums) or 1
                tally = dict(Counter(nums))
                result.update({
                    'avg': round(sum(nums) / len(nums), 1) if nums else None,
                    'rating_list': [
                        {'score': i, 'count': tally.get(i, 0),
                         'pct': round(tally.get(i, 0) / total * 100)}
                        for i in range(1, 6)
                    ],
                })

            elif q.q_type == 'yes_no':
                tally = dict(Counter(raw_answers))
                total = sum(tally.values()) or 1
                result.update({
                    'yn_list': [
                        {'label': 'Yes', 'count': tally.get('Yes', 0),
                         'pct': round(tally.get('Yes', 0) / total * 100), 'color': '#10b981'},
                        {'label': 'No', 'count': tally.get('No', 0),
                         'pct': round(tally.get('No', 0) / total * 100), 'color': '#f87171'},
                    ],
                    'total': total,
                })

            question_results.append(result)

        return render(request, 'projects/survey_detail.html', {
            'project': project,
            'survey': survey,
            'questions': questions,
            'responses': responses,
            'question_results': question_results,
            'is_manager': is_manager,
            'is_member': is_member,
            'already_responded': already_responded,
        })


class SurveySubmitView(LoginRequiredMixin, View):
    """Survey fill-in page for any authenticated user."""

    def _guard(self, request, survey, pk):
        if not survey.is_active:
            messages.warning(request, 'This survey is no longer accepting responses.')
            return redirect('project_surveys', pk=pk)

        if not survey.is_anonymous and survey.responses.filter(respondent=request.user).exists():
            messages.info(request, 'You have already submitted a response to this survey.')
            return redirect('survey_dashboard', pk=pk, sid=survey.pk)

        return None

    def get(self, request, pk, sid):
        project = get_object_or_404(Project, pk=pk)
        survey = get_object_or_404(Survey, pk=sid, project=project)
        if (guard := self._guard(request, survey, pk)):
            return guard

        return render(request, 'projects/survey_submit.html', {
            'project': project,
            'survey': survey,
            'questions': survey.questions.order_by('order'),
        })

    def post(self, request, pk, sid):
        project = get_object_or_404(Project, pk=pk)
        survey = get_object_or_404(Survey, pk=sid, project=project)
        if (guard := self._guard(request, survey, pk)):
            return guard

        questions = list(survey.questions.order_by('order'))

        errors = []
        cleaned_answers = []

        for q in questions:
            field_name = f'q_{q.pk}'

            if q.q_type == 'multi_choice':
                vals = [v.strip() for v in request.POST.getlist(field_name) if v.strip()]
                val = ', '.join(vals)
            else:
                val = request.POST.get(field_name, '').strip()

            if q.is_required and not val:
                errors.append(f'Please answer: {q.text}')

            cleaned_answers.append((q, val))

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'projects/survey_submit.html', {
                'project': project,
                'survey': survey,
                'questions': questions,
            })

        resp = SurveyResponse.objects.create(
            survey=survey,
            respondent=None if survey.is_anonymous else request.user,
            respondent_name=(
                request.POST.get('respondent_name', '').strip()
                if survey.is_anonymous
                else request.user.get_full_name()
            ),
        )

        for q, val in cleaned_answers:
            SurveyAnswer.objects.create(
                response=resp,
                question=q,
                value=val
            )

        messages.success(request, 'Response submitted — thank you!')
        return redirect('survey_dashboard', pk=pk, sid=survey.pk)


class SurveyToggleView(LoginRequiredMixin, View):
    """Activate / close a survey, or toggle public access (project manager only)."""

    def post(self, request, pk, sid):
        project = get_object_or_404(Project, pk=pk)
        if not (request.user.is_superuser or project.project_manager == request.user):
            return HttpResponseForbidden()
        survey = get_object_or_404(Survey, pk=sid, project=project)

        action = request.POST.get('action', 'toggle_active')

        if action == 'toggle_public':
            survey.is_public = not survey.is_public
            survey.allow_multiple_responses = request.POST.get('allow_multiple') == 'on'
            survey.save(update_fields=['is_public', 'allow_multiple_responses'])
            label = 'open to the public' if survey.is_public else 'restricted to internal users'
            messages.success(request, f'Survey "{survey.title}" is now {label}.')
        else:
            survey.is_active = not survey.is_active
            survey.save(update_fields=['is_active'])
            label = 'activated' if survey.is_active else 'closed'
            messages.success(request, f'Survey "{survey.title}" {label}.')

        return redirect('survey_detail', pk=pk, sid=sid)


class SurveyDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk, sid):
        project = get_object_or_404(Project, pk=pk)
        if not (request.user.is_superuser or project.project_manager == request.user):
            return HttpResponseForbidden()
        survey = get_object_or_404(Survey, pk=sid, project=project)
        survey.delete()
        messages.success(request, 'Survey deleted.')
        return redirect('project_surveys', pk=pk)

class SurveyEditView(LoginRequiredMixin, View):
    """Edit survey settings and rebuild its questions."""

    def post(self, request, pk, sid):
        project = get_object_or_404(Project, pk=pk)
        if not (request.user.is_superuser or project.project_manager == request.user):
            return HttpResponseForbidden()

        survey = get_object_or_404(Survey, pk=sid, project=project)

        title = request.POST.get('title', '').strip()
        if not title:
            messages.error(request, 'Survey title is required.')
            return redirect('survey_detail', pk=pk, sid=sid)

        survey.title = title
        survey.description = request.POST.get('description', '').strip()
        survey.is_anonymous = request.POST.get('is_anonymous') == 'on'
        survey.is_public = request.POST.get('is_public') == 'on'
        survey.allow_multiple_responses = request.POST.get('allow_multiple_responses') == 'on'
        survey.closes_at = request.POST.get('closes_at') or None

        if survey.is_public and not survey.public_token:
            import uuid
            survey.public_token = uuid.uuid4()

        survey.save()

        _save_survey_questions(survey, request)

        messages.success(request, f'Survey "{survey.title}" updated.')
        return redirect('survey_dashboard', pk=pk, sid=sid)

class TrackingSheetView(LoginRequiredMixin, View):
    """Main tracking sheet (get-or-create per project)."""

    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        sheet, _ = TrackingSheet.objects.get_or_create(
            project=project,
            defaults={'title': f'{project.name} Tracker', 'created_by': request.user}
        )
        columns = sheet.get_columns()
        rows_qs = sheet.rows.select_related('created_by').order_by('order', 'created_at')

        rows_processed = []
        for row in rows_qs:
            data = row.get_data()
            cells = [
                (col['label'], col['type'], col['key'], data.get(col['key'], ''))
                for col in columns
            ]
            rows_processed.append({
                'row': row,
                'cells': cells,
                'data_json': json.dumps(data),
            })

        return render(request, 'projects/tracking_sheet.html', {
            'project': project,
            'sheet': sheet,
            'columns': columns,
            'rows_processed': rows_processed,
            'row_count': len(rows_processed),
            'is_member': _is_project_member(request.user, project),
            'is_manager': request.user.is_superuser or project.project_manager == request.user,
        })


class TrackingSheetUpdateColumnsView(LoginRequiredMixin, View):
    """Save column definitions (manager only)."""

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not (request.user.is_superuser or project.project_manager == request.user):
            return HttpResponseForbidden()

        sheet, _ = TrackingSheet.objects.get_or_create(
            project=project,
            defaults={'created_by': request.user}
        )

        sheet.title = request.POST.get('sheet_title', sheet.title).strip() or sheet.title

        labels = request.POST.getlist('col_label')
        types = request.POST.getlist('col_type')
        keys = request.POST.getlist('col_key')  # preserve existing keys

        cols = []
        for i, (label, ctype) in enumerate(zip(labels, types)):
            label = (label or '').strip()
            if not label:
                continue

            old_key = ''
            if i < len(keys):
                old_key = (keys[i] or '').strip()

            if old_key:
                key = old_key
            else:
                safe = label.lower().replace(' ', '_')
                key = f'c{i}_{safe}'

            cols.append({
                'key': key,
                'label': label,
                'type': ctype,
            })

        sheet.set_columns(cols)
        sheet.save()

        messages.success(request, 'Columns updated.')
        return redirect('project_tracking', pk=pk)


class TrackingRowAddView(LoginRequiredMixin, View):
    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()
        sheet, _ = TrackingSheet.objects.get_or_create(
            project=project, defaults={'created_by': request.user}
        )
        data = {col['key']: request.POST.get(f'data_{col["key"]}', '') for col in sheet.get_columns()}
        TrackingRow.objects.create(
            sheet=sheet,
            data_json=json.dumps(data),
            order=sheet.rows.count(),
            created_by=request.user,
        )
        messages.success(request, 'Row added.')
        return redirect('project_tracking', pk=pk)


class TrackingRowEditView(LoginRequiredMixin, View):
    def post(self, request, pk, rid):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        row = get_object_or_404(TrackingRow, pk=rid, sheet__project=project)

        data = {}
        for col in row.sheet.get_columns():
            field_name = f'data_{col["key"]}'

            if col['type'] == 'checkbox':
                data[col['key']] = request.POST.get(field_name) == 'true'
            else:
                data[col['key']] = request.POST.get(field_name, '')

        row.data_json = json.dumps(data)
        row.save(update_fields=['data_json'])

        messages.success(request, 'Row updated.')
        return redirect('project_tracking', pk=pk)


class TrackingRowDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk, rid):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()
        get_object_or_404(TrackingRow, pk=rid, sheet__project=project).delete()
        messages.success(request, 'Row deleted.')
        return redirect('project_tracking', pk=pk)


class TrackingSheetExportCSVView(LoginRequiredMixin, View):
    def get(self, request, pk):
        import csv as _csv
        project = get_object_or_404(Project, pk=pk)
        sheet = get_object_or_404(TrackingSheet, project=project)
        columns = sheet.get_columns()
        rows = sheet.rows.select_related('created_by').order_by('order', 'created_at')

        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{project.code}_tracking.csv"'
        writer = _csv.writer(resp)
        writer.writerow(['#'] + [c['label'] for c in columns] + ['Added By', 'Date'])

        for i, row in enumerate(rows, 1):
            d = row.get_data()
            writer.writerow(
                [i]
                + [d.get(c['key'], '') for c in columns]
                + [
                    row.created_by.get_full_name() if row.created_by else '',
                    row.created_at.strftime('%Y-%m-%d'),
                ]
            )
        return resp


# =============================================================================
# LOCATION MAP VIEWS
# =============================================================================

class ProjectLocationMapView(LoginRequiredMixin, View):
    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        locations = project.locations.select_related('assigned_to').order_by('name')
        from apps.projects.models import ProjectZone
        zones = project.zones.all()
        return render(request, 'projects/location_map.html', {
            'project': project,
            'locations': locations,
            'location_count': locations.count(),
            'is_member': _is_project_member(request.user, project),
            'is_manager': request.user.is_superuser or project.project_manager == request.user,
            'status_choices': ProjectLocation.Status.choices,
            'staff_list': User.objects.filter(is_active=True).order_by('first_name'),
            'zones': zones,
            'zone_count': zones.count(),
        })


class ProjectLocationAddView(LoginRequiredMixin, View):
    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()
        try:
            lat = float(request.POST['latitude'])
            lng = float(request.POST['longitude'])
        except (KeyError, ValueError):
            messages.error(request, 'Valid latitude and longitude are required.')
            return redirect('project_locations', pk=pk)

        assigned_id = request.POST.get('assigned_to') or None

        assigned = None
        if assigned_id:
            assigned = User.objects.filter(pk=assigned_id).first()
        from apps.projects.models import ProjectZone
        zone_id = request.POST.get('zone') or None
        zone = None
        if zone_id:
            zone = ProjectZone.objects.filter(pk=zone_id, project=project).first()
        ProjectLocation.objects.create(
            project=project,
            name=request.POST.get('name', '').strip(),
            description=request.POST.get('description', '').strip(),
            address=request.POST.get('address', '').strip(),
            latitude=lat,
            longitude=lng,
            category=request.POST.get('category', '').strip(),
            status=request.POST.get('status', 'pending'),
            assigned_to=assigned,
            notes=request.POST.get('notes', '').strip(),
            zone=zone,
        )
        messages.success(request, 'Location added.')
        return redirect('project_locations', pk=pk)


class ProjectLocationEditView(LoginRequiredMixin, View):
    def post(self, request, pk, lid):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()
        loc = get_object_or_404(ProjectLocation, pk=lid, project=project)
        try:
            lat = float(request.POST['latitude'])
            lng = float(request.POST['longitude'])
        except (KeyError, ValueError):
            messages.error(request, 'Valid coordinates required.')
            return redirect('project_locations', pk=pk)

        loc.name = request.POST.get('name', '').strip()
        loc.description = request.POST.get('description', '').strip()
        loc.address = request.POST.get('address', '').strip()
        loc.latitude = lat
        loc.longitude = lng
        loc.category = request.POST.get('category', '').strip()
        loc.status = request.POST.get('status', 'pending')
        assigned_id = request.POST.get('assigned_to') or None
        loc.assigned_to = User.objects.filter(pk=assigned_id).first() if assigned_id else None
        loc.notes = request.POST.get('notes', '').strip()
        from apps.projects.models import ProjectZone
        zone_id = request.POST.get('zone') or None
        loc.zone = ProjectZone.objects.filter(pk=zone_id, project=project).first() if zone_id else None
        loc.save()
        messages.success(request, 'Location updated.')
        return redirect('project_locations', pk=pk)


class ProjectLocationDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk, lid):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()
        get_object_or_404(ProjectLocation, pk=lid, project=project).delete()
        messages.success(request, 'Location removed.')
        return redirect('project_locations', pk=pk)


class ProjectLocationImportView(LoginRequiredMixin, View):
    """Bulk-import locations from a CSV or Excel file."""

    FIELD_MAP = {
        'name':        ['name', 'location name', 'location'],
        'latitude':    ['latitude', 'lat'],
        'longitude':   ['longitude', 'lon', 'lng', 'long'],
        'description': ['description', 'desc'],
        'address':     ['address'],
        'category':    ['category', 'type'],
        'status':      ['status'],
        'notes':       ['notes', 'note'],
    }
    VALID_STATUSES = {'pending', 'in_progress', 'completed', 'on_hold'}

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        upload = request.FILES.get('import_file')
        if not upload:
            messages.error(request, 'No file selected.')
            return redirect('project_locations', pk=pk)

        fname = upload.name.lower()
        rows, parse_error = [], None

        try:
            if fname.endswith('.csv'):
                import csv, io as _io
                text = upload.read().decode('utf-8-sig')
                rows = list(csv.DictReader(_io.StringIO(text)))

            elif fname.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(upload, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        rows.append(dict(zip(headers, row)))
            else:
                messages.error(request, 'Only .csv and .xlsx files are supported.')
                return redirect('project_locations', pk=pk)
        except Exception as e:
            messages.error(request, f'Could not read file: {e}')
            return redirect('project_locations', pk=pk)

        def _get(row, field):
            """Pull a value from a row regardless of column alias used."""
            for alias in self.FIELD_MAP.get(field, [field]):
                for key, val in row.items():
                    if str(key).strip().lower() == alias:
                        return str(val).strip() if val is not None else ''
            return ''

        created, skipped, errors = 0, 0, []

        for i, row in enumerate(rows, 2):
            name = _get(row, 'name')
            if not name or name.startswith('#'):
                skipped += 1
                continue

            try:
                lat = float(_get(row, 'latitude'))
                lng = float(_get(row, 'longitude'))
            except ValueError:
                skipped += 1
                errors.append(f'Row {i} ("{name}"): invalid coordinates.')
                continue

            if not (-90 <= lat <= 90) or not (-180 <= lng <= 180):
                skipped += 1
                errors.append(f'Row {i} ("{name}"): coordinates out of range.')
                continue

            status = _get(row, 'status').lower().replace(' ', '_')
            if status not in self.VALID_STATUSES:
                status = 'pending'

            ProjectLocation.objects.create(
                project=project,
                name=name,
                description=_get(row, 'description'),
                address=_get(row, 'address'),
                latitude=lat,
                longitude=lng,
                category=_get(row, 'category'),
                status=status,
                notes=_get(row, 'notes'),
            )
            created += 1

        if errors:
            messages.warning(
                request,
                f'Imported {created} location(s); {skipped} row(s) skipped. '
                + ' | '.join(errors[:5])
                + (' …' if len(errors) > 5 else '')
            )
        else:
            messages.success(request, f'Successfully imported {created} location(s).')

        return redirect('project_locations', pk=pk)


class LocationTemplateDownloadView(LoginRequiredMixin, View):
    """Download a blank CSV or Excel template for bulk location import."""

    HEADERS  = ['name', 'description', 'address', 'latitude', 'longitude',
                 'category', 'status', 'notes']
    SAMPLE   = ['Site A – Banjul Office', 'Main coordination hub',
                 '12 Marina Parade, Banjul', '13.4549', '-16.5790',
                 'Office', 'active', 'Near the waterfront']
    SAMPLE2  = ['Field Site B', 'Installation point',
                 'Kanifing Industrial Zone', '13.3950', '-16.6100',
                 'Installation', 'pending', '']
    STATUS_NOTE = ('# status values: pending | in_progress | completed | on_hold  '
                   '— delete this row before importing')

    def get(self, request, pk, fmt='xlsx'):
        project = get_object_or_404(Project, pk=pk)

        if fmt == 'csv':
            import csv, io as _io
            buf = _io.StringIO()
            w = csv.writer(buf)
            w.writerow(self.HEADERS)
            w.writerow(self.SAMPLE)
            w.writerow(self.SAMPLE2)
            w.writerow([self.STATUS_NOTE] + [''] * (len(self.HEADERS) - 1))
            resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = (
                f'attachment; filename="{project.code}_locations_template.csv"'
            )
            return resp

        elif fmt == 'xlsx':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                import io as _io

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Locations'

                # ── Header row ──
                hdr_fill = PatternFill('solid', fgColor=project.cover_color.lstrip('#') or '1e3a5f')
                thin = Side(style='thin', color='D0D7E0')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                for col, h in enumerate(self.HEADERS, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.font      = Font(bold=True, color='FFFFFF', size=11)
                    c.fill      = hdr_fill
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border    = border
                ws.row_dimensions[1].height = 20

                # ── Sample rows ──
                for r_idx, sample in enumerate([self.SAMPLE, self.SAMPLE2], 2):
                    row_fill = PatternFill('solid', fgColor='F8FAFC' if r_idx % 2 == 0 else 'FFFFFF')
                    for col, val in enumerate(sample, 1):
                        c = ws.cell(row=r_idx, column=col, value=val)
                        c.alignment = Alignment(vertical='center')
                        c.fill   = row_fill
                        c.border = border

                # ── Notes row ──
                nc = ws.cell(row=4, column=1, value=self.STATUS_NOTE)
                nc.font = Font(italic=True, color='94A3B8', size=9)

                # ── Column widths ──
                widths = [28, 35, 35, 12, 12, 18, 14, 28]
                for col, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(col)].width = w

                # ── Freeze header ──
                ws.freeze_panes = 'A2'

                buf = _io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                resp = HttpResponse(
                    buf.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                resp['Content-Disposition'] = (
                    f'attachment; filename="{project.code}_locations_template.xlsx"'
                )
                return resp

            except ImportError:
                messages.error(request, 'openpyxl is required for Excel export. Run: pip install openpyxl')
                return redirect('project_locations', pk=pk)

        return HttpResponse('Invalid format.', status=400)


# =============================================================================
# LOCATION MAP HTML and CSV EXPORT
# =============================================================================

class LocationMapExportCSVView(LoginRequiredMixin, View):
    """
    GET: Returns a CSV with geo-friendly columns. Works for direct download
    AND for the "Save to Files" flow (if ?save_to_files=1&format=csv|xlsx is
    supplied, the file is written as a SharedFile in the user's files and we
    JSON-redirect back to the files app).
    """

    HEADERS = [
        'ID', 'Name', 'Description', 'Address',
        'Latitude', 'Longitude',
        'Accuracy (m)', 'Source',
        'Category', 'Status', 'Assigned To',
        'Zone', 'Notes', 'Created At',
    ]

    def _row(self, loc):
        return [
            str(loc.pk),
            loc.name,
            loc.description or '',
            loc.address or '',
            f'{loc.latitude:.6f}' if loc.latitude is not None else '',
            f'{loc.longitude:.6f}' if loc.longitude is not None else '',
            '',  # Accuracy — ProjectLocation has no accuracy field
            'manual',  # Source — all manual unless imported
            loc.category or '',
            loc.get_status_display(),
            loc.assigned_to.get_full_name() if loc.assigned_to else '',
            loc.zone.name if loc.zone else '',
            loc.notes or '',
            loc.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        ]

    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        fmt = (request.GET.get('format') or 'csv').lower()
        save_to_files = request.GET.get('save_to_files') == '1'

        locations = project.locations.select_related('assigned_to', 'zone').all()

        if fmt == 'xlsx':
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Locations'
            ws.append(self.HEADERS)
            for loc in locations:
                ws.append(self._row(loc))
            # Freeze header row + widen some columns
            ws.freeze_panes = 'A2'
            for i, w in enumerate([32, 28, 40, 40, 14, 14, 14, 14, 18, 14, 20, 16, 30, 20], start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            content = buf.read()
            mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            out_name = f'{project.code}_locations.xlsx'
        else:
            sio = io.StringIO()
            # utf-8 BOM so Excel opens it correctly
            sio.write('\ufeff')
            writer = csv.writer(sio)
            writer.writerow(self.HEADERS)
            for loc in locations:
                writer.writerow(self._row(loc))
            content = sio.getvalue().encode('utf-8')
            mime = 'text/csv; charset=utf-8'
            out_name = f'{project.code}_locations.csv'

        # ── Save to Files app as a SharedFile ────────────────────────────────
        if save_to_files:
            sf = SharedFile.objects.create(
                name=out_name,
                file=ContentFile(content, name=out_name),
                project=project,
                uploaded_by=request.user,
                visibility='private',
                description=f'Location map export from project {project.code}',
                tags='export,locations',
                file_size=len(content),
                file_type=mime.split(';')[0].strip(),
            )
            try:
                sf.file_hash = sf.compute_hash()
                sf.save(update_fields=['file_hash'])
            except Exception:
                pass
            return JsonResponse({
                'ok': True,
                'message': f'Saved "{out_name}" to your Files.',
                'file_id': str(sf.pk),
                'file_name': sf.name,
            })

        # ── Direct download ──────────────────────────────────────────────────
        resp = HttpResponse(content, content_type=mime)
        resp['Content-Disposition'] = f'attachment; filename="{out_name}"'
        return resp

class LocationMapHTMLExportView(LoginRequiredMixin, View):
    """Download a self-contained interactive Leaflet map as an HTML file."""

    _STATUS_COLORS = {
        'pending': '#94a3b8',
        'in_progress': '#3b82f6',
        'completed': '#10b981',
        'on_hold': '#f59e0b',
    }

    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        locations = project.locations.select_related('assigned_to').order_by('name')

        # Convert lazy labels to plain strings
        status_labels = {
            value: str(label)
            for value, label in ProjectLocation.Status.choices
        }

        loc_data = [
            {
                'name': str(loc.name or ''),
                'description': str(loc.description or ''),
                'address': str(loc.address or ''),
                'lat': float(loc.latitude),
                'lng': float(loc.longitude),
                'category': str(loc.category or ''),
                'status': str(loc.status or ''),
                'statusLabel': str(status_labels.get(loc.status, loc.status)),
                'color': self._STATUS_COLORS.get(loc.status, '#64748b'),
                'assignedTo': str(loc.assigned_to.get_full_name() if loc.assigned_to else ''),
                'notes': str(loc.notes or ''),
            }
            for loc in locations
        ]

        html = self._render_html(project, loc_data)
        resp = HttpResponse(html, content_type='text/html; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{project.code}_location_map.html"'
        return resp

    def _render_html(self, project, loc_data):
        import json as _json
        from django.utils.timezone import now
        from django.utils.html import escape

        count = len(loc_data)
        exported = now().strftime('%d %b %Y')

        project_name = escape(str(project.name))
        project_code = escape(str(project.code))
        cover_color = escape(str(project.cover_color or '#1e3a5f'))

        return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{project_name} — Location Map</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" crossorigin=""/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js" crossorigin=""></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:system-ui,-apple-system,sans-serif;background:#f8fafc;display:flex;flex-direction:column;height:100vh}}
  header{{background:{cover_color};color:#fff;padding:14px 22px;display:flex;align-items:center;justify-content:space-between;flex-shrink:0}}
  header h1{{font-size:1.1rem;font-weight:700}}
  header p{{font-size:.82rem;opacity:.92;margin-top:2px}}
  .wrap{{display:grid;grid-template-columns:340px 1fr;min-height:0;flex:1}}
  .sidebar{{background:#fff;border-right:1px solid #e2e8f0;overflow:auto}}
  .sidebar-head{{padding:14px 16px;border-bottom:1px solid #e2e8f0}}
  .sidebar-head h2{{font-size:.92rem;font-weight:700;color:#0f172a}}
  .sidebar-head p{{font-size:.76rem;color:#64748b;margin-top:4px}}
  .list{{padding:12px;display:flex;flex-direction:column;gap:10px}}
  .card{{border:1px solid #e2e8f0;border-radius:12px;padding:12px;cursor:pointer;transition:.15s}}
  .card:hover{{border-color:#3b82f6;box-shadow:0 3px 14px rgba(0,0,0,.06)}}
  .card-title{{font-size:.88rem;font-weight:700;color:#0f172a}}
  .card-meta{{font-size:.75rem;color:#64748b;margin-top:4px}}
  .badge{{display:inline-block;padding:3px 8px;border-radius:999px;font-size:.68rem;font-weight:700;margin-top:8px}}
  #map{{height:100%;width:100%}}
  @media (max-width: 900px) {{
    .wrap{{grid-template-columns:1fr}}
    .sidebar{{max-height:40vh}}
  }}
</style>
</head>
<body>
<header>
  <div>
    <h1>{project_name} — Location Map</h1>
    <p>{count} location{"s" if count != 1 else ""} • Exported {exported}</p>
  </div>
  <div style="font-size:.8rem;opacity:.9">{project_code}</div>
</header>

<div class="wrap">
  <aside class="sidebar">
    <div class="sidebar-head">
      <h2>Locations</h2>
      <p>Click a location to focus it on the map.</p>
    </div>
    <div class="list" id="loc-list"></div>
  </aside>
  <main id="map"></main>
</div>

<script>
var LOCS = {_json.dumps(loc_data, ensure_ascii=False)};
var STATUS_COLORS = {{
  pending: '#94a3b8',
  in_progress: '#3b82f6',
  completed: '#10b981',
  on_hold: '#f59e0b'
}};

function makeIcon(color) {{
  var svg = '<svg xmlns="http://www.w3.org/2000/svg" width="28" height="36" viewBox="0 0 28 36">'
    + '<path d="M14 0C6.3 0 0 6.3 0 14c0 9.3 14 22 14 22S28 23.3 28 14C28 6.3 21.7 0 14 0z" fill="'+color+'" stroke="#fff" stroke-width="1.5"/>'
    + '<circle cx="14" cy="14" r="5.5" fill="white"/></svg>';
  return L.divIcon({{html:svg, className:'', iconSize:[28,36], iconAnchor:[14,36], popupAnchor:[0,-38]}});
}}

var map = L.map('map');
L.tileLayer('https://{{s}}.basemaps.cartocdn.com/light_all/{{z}}/{{x}}/{{y}}{{r}}.png', {{
  attribution: '&copy; OpenStreetMap contributors &copy; CARTO',
  subdomains: 'abcd',
  maxZoom: 20
}}).addTo(map);

var bounds = [];
var markers = [];
var listEl = document.getElementById('loc-list');

LOCS.forEach(function(loc, idx) {{
  var color = loc.color || STATUS_COLORS[loc.status] || '#64748b';

  var popup = '<div style="font-weight:700;font-size:.92rem;margin-bottom:6px">' + (loc.name || '') + '</div>'
    + '<span style="font-size:.7rem;font-weight:700;padding:2px 7px;border-radius:999px;background:'+color+'20;color:'+color+'">' + (loc.statusLabel || '') + '</span>';

  if (loc.category) popup += '<div style="font-size:.78rem;color:#475569;margin-top:6px"><b>Category:</b> ' + loc.category + '</div>';
  if (loc.address) popup += '<div style="font-size:.78rem;color:#475569"><b>Address:</b> ' + loc.address + '</div>';
  if (loc.assignedTo) popup += '<div style="font-size:.78rem;color:#475569"><b>Assigned:</b> ' + loc.assignedTo + '</div>';
  if (loc.description) popup += '<div style="font-size:.78rem;color:#475569;margin-top:4px">' + loc.description + '</div>';
  if (loc.notes) popup += '<div style="font-size:.78rem;color:#475569;margin-top:4px"><b>Notes:</b> ' + loc.notes + '</div>';
  popup += '<div style="font-size:.72rem;color:#94a3b8;margin-top:6px;font-family:monospace">' + loc.lat + ', ' + loc.lng + '</div>';

  var marker = L.marker([loc.lat, loc.lng], {{icon: makeIcon(color)}}).addTo(map);
  marker.bindPopup(popup, {{maxWidth: 280}});
  markers.push(marker);
  bounds.push([loc.lat, loc.lng]);

  var badgeBg = color + '20';
  var card = document.createElement('div');
  card.className = 'card';
  card.innerHTML =
      '<div class="card-title">' + (loc.name || '') + '</div>'
    + (loc.category ? '<div class="card-meta">' + loc.category + '</div>' : '')
    + '<span class="badge" style="background:'+badgeBg+';color:'+color+'">' + (loc.statusLabel || '') + '</span>'
    + (loc.address ? '<div class="card-meta">' + loc.address + '</div>' : '');

  card.addEventListener('click', function() {{
    map.flyTo([loc.lat, loc.lng], 16, {{duration: 0.8}});
    setTimeout(function() {{ marker.openPopup(); }}, 500);
  }});

  listEl.appendChild(card);
}});

if (bounds.length > 1) {{
  map.fitBounds(bounds, {{padding:[40,40]}});
}} else if (bounds.length === 1) {{
  map.setView(bounds[0], 15);
}} else {{
  map.setView([13.4549, -16.5790], 8);
}}
</script>
</body>
</html>"""


# =============================================================================
# LOCATION MAP PDF EXPORT
# =============================================================================

class LocationMapPDFExportView(LoginRequiredMixin, View):
    """Export a formatted PDF table of all project locations."""

    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        locations = list(project.locations.select_related('assigned_to').order_by('name'))

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors as RC
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            import io as _io

            buf = _io.BytesIO()
            doc = SimpleDocTemplate(
                buf, pagesize=landscape(A4),
                topMargin=16 * mm, bottomMargin=16 * mm,
                leftMargin=14 * mm, rightMargin=14 * mm,
            )

            hex_ = project.cover_color.lstrip('#')
            C_COVER = RC.Color(*[int(hex_[i:i + 2], 16) / 255 for i in (0, 2, 4)])

            S = getSampleStyleSheet()
            title_style = ParagraphStyle('T', fontSize=16, fontName='Helvetica-Bold',
                                         textColor=C_COVER, spaceAfter=3)
            sub_style = ParagraphStyle('S', fontSize=9, fontName='Helvetica',
                                       textColor=RC.HexColor('#64748b'), spaceAfter=14)
            note_style = ParagraphStyle('N', fontSize=8, fontName='Helvetica',
                                        textColor=RC.HexColor('#475569'), spaceAfter=5)

            STATUS_C = {
                'pending': RC.HexColor('#94a3b8'),
                'in_progress': RC.HexColor('#3b82f6'),
                'completed': RC.HexColor('#10b981'),
                'on_hold': RC.HexColor('#f59e0b'),
            }
            STATUS_L = dict(ProjectLocation.Status.choices)

            from django.utils.timezone import now
            story = [
                Paragraph(f'{project.name} — Location Map', title_style),
                Paragraph(
                    f'{project.code} · {len(locations)} location(s) · Exported {now().strftime("%d %b %Y")}',
                    sub_style
                ),
            ]

            # Table data
            headers = ['#', 'Location Name', 'Category', 'Address', 'Lat', 'Long', 'Status', 'Assigned To']
            rows = [headers]
            for i, loc in enumerate(locations, 1):
                rows.append([
                    str(i),
                    loc.name,
                    loc.category or '—',
                    loc.address or '—',
                    f'{float(loc.latitude):.5f}',
                    f'{float(loc.longitude):.5f}',
                    STATUS_L.get(loc.status, loc.status),
                    loc.assigned_to.get_full_name() if loc.assigned_to else '—',
                ])

            col_w = [8 * mm, 50 * mm, 32 * mm, 55 * mm, 22 * mm, 22 * mm, 24 * mm, 34 * mm]
            tbl = Table(rows, colWidths=col_w, repeatRows=1)
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), C_COVER),
                ('TEXTCOLOR', (0, 0), (-1, 0), RC.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7.5),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [RC.white, RC.HexColor('#f8fafc')]),
                ('GRID', (0, 0), (-1, -1), 0.3, RC.HexColor('#e2e8f0')),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (4, 1), (5, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ])
            # Colour status column per row
            for ri, loc in enumerate(locations, 1):
                sc = STATUS_C.get(loc.status, RC.HexColor('#64748b'))
                style.add('TEXTCOLOR', (6, ri), (6, ri), sc)
                style.add('FONTNAME', (6, ri), (6, ri), 'Helvetica-Bold')
            tbl.setStyle(style)
            story.append(tbl)

            # Notes
            locs_with_notes = [(l, l.notes) for l in locations if l.notes.strip()]
            if locs_with_notes:
                story += [Spacer(1, 14),
                          Paragraph('Notes', ParagraphStyle('H2', fontSize=10, fontName='Helvetica-Bold',
                                                            textColor=C_COVER, spaceAfter=6))]
                for loc, note in locs_with_notes:
                    story.append(Paragraph(f'<b>{loc.name}:</b> {note}', note_style))

            doc.build(story)
            buf.seek(0)
            return HttpResponse(
                buf.getvalue(),
                content_type='application/pdf',
                headers={'Content-Disposition': f'attachment; filename="{project.code}_locations.pdf"'},
            )

        except ImportError:
            return HttpResponse('ReportLab is required for PDF export.', status=500)

# =============================================================================
# PUBLIC SURVEY — no login required
# =============================================================================

def _get_client_ip(request):
    """Extract real IP even behind a proxy."""
    xff = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')

def _save_survey_questions(survey, request):
    """
    Rebuild survey questions from posted form fields.
    """
    survey.questions.all().delete()

    q_texts = request.POST.getlist('q_text')
    q_types = request.POST.getlist('q_type')
    q_options = request.POST.getlist('q_options')
    q_required = set(request.POST.getlist('q_required'))

    order = 0
    for i, text in enumerate(q_texts):
        text = (text or '').strip()
        if not text:
            continue

        SurveyQuestion.objects.create(
            survey=survey,
            text=text,
            q_type=q_types[i] if i < len(q_types) else 'text',
            options=(q_options[i] if i < len(q_options) else '').strip(),
            is_required=str(i) in q_required,
            order=order,
        )
        order += 1

class PublicSurveyView(View):
    """
    Token-based public survey — accessible without login.
    URL: /projects/surveys/public/<token>/
    Captures IP, User-Agent and optional JS device fingerprint.
    """

    def _guard(self, survey):
        from django.utils import timezone as tz
        if not survey.is_active:
            return 'closed'
        if not survey.is_public:
            return 'private'
        if survey.closes_at and survey.closes_at < tz.now().date():
            return 'expired'
        return None

    def _already_responded(self, request, survey):
        """Block duplicate submissions by IP when multiple responses off."""
        if survey.allow_multiple_responses:
            return False
        ip = _get_client_ip(request)
        device_id = request.COOKIES.get('_fp', '')
        qs = survey.responses.filter(is_public_response=True)
        if ip:
            qs = qs.filter(ip_address=ip)
        if device_id:
            qs = qs.filter(device_id=device_id)
        return qs.exists()

    def get(self, request, token):
        survey = get_object_or_404(Survey, public_token=token)
        reason = self._guard(survey)
        if reason:
            return render(request, 'projects/survey_public_closed.html', {
                'survey': survey, 'reason': reason
            })
        already = self._already_responded(request, survey)
        return render(request, 'projects/survey_public.html', {
            'survey': survey,
            'questions': survey.questions.order_by('order'),
            'already_responded': already,
        })

    def post(self, request, token):
        survey = get_object_or_404(Survey, public_token=token)
        reason = self._guard(survey)
        if reason:
            return render(request, 'projects/survey_public_closed.html', {
                'survey': survey, 'reason': reason
            })
        if self._already_responded(request, survey):
            return render(request, 'projects/survey_public.html', {
                'survey': survey,
                'questions': survey.questions.order_by('order'),
                'already_responded': True,
            })

        ip = _get_client_ip(request)
        device_id = request.POST.get('_fp', '') or request.COOKIES.get('_fp', '')
        user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
        name = request.POST.get('respondent_name', '').strip()

        resp = SurveyResponse.objects.create(
            survey=survey,
            respondent=None,
            respondent_name=name,
            ip_address=ip or None,
            device_id=device_id,
            user_agent=user_agent,
            is_public_response=True,
        )

        for q in survey.questions.order_by('order'):
            if q.q_type == 'multi_choice':
                val = ', '.join(request.POST.getlist(f'q_{q.pk}'))
            else:
                val = request.POST.get(f'q_{q.pk}', '').strip()
            SurveyAnswer.objects.create(response=resp, question=q, value=val)

        response = render(request, 'projects/survey_public_thanks.html', {'survey': survey})
        # Set a cookie so the browser remembers it submitted (30-day)
        response.set_cookie('_survey_done_' + str(survey.pk), '1', max_age=60 * 60 * 24 * 30, httponly=True)
        return response


# =============================================================================
# SURVEY DASHBOARD  — live charts for CEO / Admin / PM
# =============================================================================

def _build_survey_chart_data(survey):
    """
    Returns a JSON-serialisable payload for the survey live dashboard.
    Supports:
      - single_choice / multi_choice
      - rating
      - yes_no
      - text / paragraph / number
      - geolocation
    """
    from collections import Counter, defaultdict
    from django.utils import timezone as tz
    from django.utils.timezone import localtime

    questions = list(survey.questions.order_by('order'))
    responses = list(
        survey.responses
        .select_related('respondent')
        .prefetch_related('answers__question')
        .order_by('submitted_at')
    )
    total = len(responses)

    # Map answers by question
    answers_by_q = {str(q.pk): [] for q in questions}
    for resp in responses:
        for ans in resp.answers.all():
            answers_by_q.setdefault(str(ans.question_id), []).append(ans.value or '')

    question_charts = []

    for q in questions:
        raw = answers_by_q.get(str(q.pk), [])
        entry = {
            'id': str(q.pk),
            'text': q.text,
            'type': q.q_type,
            'count': len([v for v in raw if str(v).strip()]),
        }

        if q.q_type in ('single_choice', 'multi_choice'):
            all_vals = []
            for a in raw:
                all_vals.extend([v.strip() for v in a.split(',') if v.strip()])
            tally = Counter(all_vals)
            opts = q.get_options_list()
            labels, counts = [], []

            for opt in opts:
                labels.append(opt)
                counts.append(tally.get(opt, 0))

            for k, v in tally.items():
                if k not in opts:
                    labels.append(k + ' *')
                    counts.append(v)

            entry.update({
                'labels': labels,
                'counts': counts,
                'total': sum(tally.values()),
            })

        elif q.q_type == 'rating':
            nums = [int(a) for a in raw if str(a).strip().isdigit()]
            tally = Counter(nums)
            entry.update({
                'labels': ['1', '2', '3', '4', '5'],
                'counts': [tally.get(i, 0) for i in range(1, 6)],
                'avg': round(sum(nums) / len(nums), 2) if nums else None,
            })

        elif q.q_type == 'yes_no':
            tally = Counter(raw)
            entry.update({
                'labels': ['Yes', 'No'],
                'counts': [tally.get('Yes', 0), tally.get('No', 0)],
                'total': sum(tally.values()) or 0,
            })

        elif q.q_type in ('text', 'paragraph', 'number'):
            entry['samples'] = [v for v in raw if str(v).strip()][:20]

        elif q.q_type == 'geolocation':
            points = []
            answer_rows = (
                SurveyAnswer.objects
                .filter(question=q)
                .select_related('response', 'response__respondent')
                .order_by('-response__submitted_at')[:500]
            )

            for ans in answer_rows:
                parsed = parse_geo_answer(ans.value)
                if not parsed:
                    continue

                resp = ans.response
                resp_name = (
                    resp.respondent_name
                    or (resp.respondent.get_full_name() if resp.respondent else 'Anonymous')
                )

                points.append({
                    'lat': parsed['lat'],
                    'lng': parsed['lng'],
                    'accuracy_m': parsed.get('accuracy_m', 0),
                    'source': parsed.get('source', 'gps'),
                    'address': parsed.get('address', ''),
                    'respondent_name': resp_name,
                    'submitted_at': localtime(resp.submitted_at).strftime('%d %b %Y %H:%M'),
                    'is_public': resp.is_public_response,
                })

            entry['points'] = points
            entry['valid_count'] = len(points)

        question_charts.append(entry)

    # Timeline
    day_counts = defaultdict(int)
    for r in responses:
        day = localtime(r.submitted_at).strftime('%Y-%m-%d')
        day_counts[day] += 1

    timeline_labels = sorted(day_counts.keys())
    timeline_counts = [day_counts[d] for d in timeline_labels]

    # Internal/public split
    internal = sum(1 for r in responses if not r.is_public_response)
    public = sum(1 for r in responses if r.is_public_response)

    # Recent responses
    recent = []
    for r in reversed(responses[-20:]):
        recent.append({
            'id': str(r.pk),
            'name': r.respondent_name or (r.respondent.get_full_name() if r.respondent else 'Anonymous'),
            'ip': r.ip_address or '—',
            'device': (r.device_id[:12] + '…') if r.device_id and len(r.device_id) > 12 else (r.device_id or '—'),
            'is_public': r.is_public_response,
            'submitted_at': localtime(r.submitted_at).strftime('%d %b %Y %H:%M'),
        })

    return {
        'survey_id': str(survey.pk),
        'title': survey.title,
        'total_responses': total,
        'internal_responses': internal,
        'public_responses': public,
        'is_active': survey.is_active,
        'question_charts': question_charts,
        'timeline': {
            'labels': timeline_labels,
            'counts': timeline_counts,
        },
        'recent': recent,
        'generated_at': tz.now().isoformat(),
    }

class SurveyDashboardDataView(LoginRequiredMixin, View):
    """
    JSON polling endpoint.  GET returns fresh chart data every ~10 s.
    Access: project manager, superuser, or team member.
    """

    def get(self, request, pk, sid):
        project = get_object_or_404(Project, pk=pk)
        if not (request.user.is_superuser or
                project.project_manager == request.user or
                _is_project_member(request.user, project)):
            return JsonResponse({'error': 'Forbidden'}, status=403)

        survey = get_object_or_404(Survey, pk=sid, project=project)
        return JsonResponse(_build_survey_chart_data(survey))


class SurveyDashboardView(LoginRequiredMixin, View):
    """
    Full-page live dashboard.
    Renders a Chart.js dashboard that polls /dashboard/data/ every 10 s.
    """

    def get(self, request, pk, sid):
        project = get_object_or_404(Project, pk=pk)
        if not (request.user.is_superuser or
                project.project_manager == request.user or
                _is_project_member(request.user, project)):
            return HttpResponseForbidden()

        survey = get_object_or_404(Survey, pk=sid, project=project)
        initial_data = _build_survey_chart_data(survey)

        import json as _json
        return render(request, 'projects/survey_dashboard.html', {
            'project': project,
            'survey': survey,
            'initial_data_json': _json.dumps(initial_data),
            'is_manager': request.user.is_superuser or project.project_manager == request.user,
            'public_url': request.build_absolute_uri(survey.get_public_url()) if survey.is_public else None,
        })


def _olc_decode(code):
    """
    Pure-Python Open Location Code (Plus Code) decoder.
    Returns (latitude_center, longitude_center).
    Mirrors the OLC JS implementation embedded in location_map.html.
    """
    ALPHABET = '23456789CFGHJMPQRVWX'
    BASE = 20
    SEP = '+'
    PAIR_LEN = 10
    GRID_ROWS = 5
    GRID_COLS = 4

    char_idx = {c: i for i, c in enumerate(ALPHABET)}
    code = code.strip().upper()

    if SEP not in code:
        raise ValueError('Missing "+" separator')

    clean = code.replace(SEP, '')
    while clean and clean[-1] == '0':
        clean = clean[:-1]

    lat, lng = -90.0, -180.0
    lat_step = 180.0 / BASE
    lng_step = 360.0 / BASE

    for i in range(0, min(len(clean), PAIR_LEN), 2):
        if clean[i] not in char_idx or clean[i + 1] not in char_idx:
            raise ValueError(f'Invalid character in Plus Code: {code!r}')
        lat += char_idx[clean[i]] * lat_step
        lng += char_idx[clean[i + 1]] * lng_step
        lat_step /= BASE
        lng_step /= BASE

    if len(clean) > PAIR_LEN:
        lgs, lns = lat_step, lng_step
        for j in range(PAIR_LEN, len(clean)):
            lgs /= GRID_ROWS
            lns /= GRID_COLS
            v = char_idx[clean[j]]
            lat += (v // GRID_COLS) * lgs
            lng += (v % GRID_COLS) * lns
        lat_step, lng_step = lgs, lns

    return lat + lat_step / 2, lng + lng_step / 2


def _olc_encode(lat, lng, length=8):
    """Encode lat/lng to a full Plus Code (used when recovering short codes)."""
    ALPHABET = '23456789CFGHJMPQRVWX'
    BASE = 20
    lat = min(90 - 1e-10, max(-90.0, lat)) + 90
    lng = ((lng + 180) % 360 + 360) % 360
    lat_step = 180.0 / BASE
    lng_step = 360.0 / BASE
    code = ''
    for _ in range(0, length, 2):
        code += ALPHABET[int(lat / lat_step)]
        lat %= lat_step
        code += ALPHABET[int(lng / lng_step)]
        lng %= lng_step
        lat_step /= BASE
        lng_step /= BASE
    return code[:8] + '+' + code[8:]


def _olc_recover(short_code, ref_lat, ref_lng):
    """
    Recover a full Plus Code from a short code using a reference lat/lng,
    then decode to (lat_center, lng_center).
    """
    SEP_POS = 8
    sep_idx = short_code.index('+')
    prefix_len = SEP_POS - sep_idx
    prefix = _olc_encode(ref_lat, ref_lng).replace('+', '')[:prefix_len]
    sn = short_code.upper().replace('+', '')
    full_clean = prefix + sn
    full_code = full_clean[:SEP_POS] + '+' + full_clean[SEP_POS:]
    return _olc_decode(full_code)


def _is_full_olc(code):
    code = code.strip().upper()
    return '+' in code and code.index('+') == 8


class PlusCodeConvertView(LoginRequiredMixin, View):
    """
    POST: accept a CSV or Excel file whose rows contain a Plus Code column.
    Decode every Plus Code to latitude / longitude, then stream the enriched
    file back to the browser.  Nothing is saved to disk on the server — the
    uploaded bytes are processed in memory and discarded immediately.

    Short codes (e.g. "C7PM+F8G") are resolved against the project's first
    saved location or a hard-coded Banjul fallback.
    """

    # Column aliases that are recognised as a Plus Code column (case-insensitive)
    PC_ALIASES = {
        'plus_code', 'plus code', 'pluscode', 'google plus code',
        'open location code', 'olc', 'plus_codes', 'location code',
    }

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        upload = request.FILES.get('pc_file')
        if not upload:
            return JsonResponse({'error': 'No file uploaded.'}, status=400)

        fname = upload.name.lower()

        # ── reference point for short-code resolution ─────────────────────────
        first_loc = project.locations.order_by('created_at').first()
        ref_lat = float(first_loc.latitude) if first_loc else 13.4549
        ref_lng = float(first_loc.longitude) if first_loc else -16.5790

        # ── parse input ───────────────────────────────────────────────────────
        try:
            if fname.endswith('.csv'):
                import csv as _csv, io as _io
                text = upload.read().decode('utf-8-sig')
                reader = _csv.DictReader(_io.StringIO(text))
                headers = reader.fieldnames or []
                rows = list(reader)
                fmt = 'csv'

            elif fname.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(upload, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        rows.append(dict(zip(headers, [
                            str(v).strip() if v is not None else '' for v in row
                        ])))
                fmt = 'xlsx'
            else:
                return JsonResponse(
                    {'error': 'Only .csv and .xlsx files are supported.'},
                    status=400
                )
        except Exception as exc:
            return JsonResponse({'error': f'Could not read file: {exc}'}, status=400)

        # ── find the Plus Code column ─────────────────────────────────────────
        pc_col = None
        for h in headers:
            if h.strip().lower() in self.PC_ALIASES:
                pc_col = h
                break

        if pc_col is None:
            return JsonResponse(
                {
                    'error': (
                        'No Plus Code column found. '
                        'Name your column one of: plus_code, pluscode, '
                        '"plus code", "google plus code", olc.'
                    )
                },
                status=400,
            )

        # ── decode ────────────────────────────────────────────────────────────
        out_headers = list(headers)
        if 'latitude' not in [h.lower() for h in out_headers]:
            out_headers.append('latitude')
        if 'longitude' not in [h.lower() for h in out_headers]:
            out_headers.append('longitude')
        # keep a "decode_error" column only if there are errors (added lazily)
        has_errors = False

        processed = []
        for row in rows:
            code = str(row.get(pc_col, '') or '').strip().upper()
            row_out = dict(row)  # copy
            if not code:
                row_out['latitude'] = ''
                row_out['longitude'] = ''
            else:
                try:
                    if _is_full_olc(code):
                        lat, lng = _olc_decode(code)
                    elif '+' in code:
                        lat, lng = _olc_recover(code, ref_lat, ref_lng)
                    else:
                        raise ValueError('Not a recognisable Plus Code')
                    row_out['latitude'] = round(lat, 6)
                    row_out['longitude'] = round(lng, 6)
                    row_out.pop('decode_error', None)
                except Exception as exc:
                    row_out['latitude'] = ''
                    row_out['longitude'] = ''
                    row_out['decode_error'] = str(exc)
                    has_errors = True

            processed.append(row_out)

        if has_errors and 'decode_error' not in out_headers:
            out_headers.append('decode_error')

        # ── stream the result back ────────────────────────────────────────────
        stem = upload.name.rsplit('.', 1)[0]

        if fmt == 'csv':
            import csv as _csv, io as _io
            buf = _io.StringIO()
            w = _csv.DictWriter(buf, fieldnames=out_headers, extrasaction='ignore')
            w.writeheader()
            w.writerows(processed)
            response = HttpResponse(
                buf.getvalue().encode('utf-8-sig'),
                content_type='text/csv; charset=utf-8',
            )
            response['Content-Disposition'] = (
                f'attachment; filename="{stem}_converted.csv"'
            )
            return response

        else:  # xlsx
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                import io as _io

                wb2 = openpyxl.Workbook()
                ws2 = wb2.active
                ws2.title = 'Converted'

                hdr_fill = PatternFill('solid', fgColor=project.cover_color.lstrip('#') or '1e3a5f')
                for col_idx, h in enumerate(out_headers, 1):
                    c = ws2.cell(row=1, column=col_idx, value=h)
                    c.font = Font(bold=True, color='FFFFFF', size=11)
                    c.fill = hdr_fill
                    c.alignment = Alignment(horizontal='center', vertical='center')
                ws2.row_dimensions[1].height = 20

                for r_idx, row in enumerate(processed, 2):
                    row_fill = PatternFill('solid', fgColor='F8FAFC' if r_idx % 2 == 0 else 'FFFFFF')
                    for c_idx, h in enumerate(out_headers, 1):
                        cell = ws2.cell(row=r_idx, column=c_idx, value=row.get(h, ''))
                        cell.fill = row_fill
                        cell.alignment = Alignment(vertical='center')

                # Auto-size columns
                for c_idx, h in enumerate(out_headers, 1):
                    ws2.column_dimensions[get_column_letter(c_idx)].width = max(
                        14, len(str(h)) + 4
                    )
                ws2.freeze_panes = 'A2'

                buf2 = _io.BytesIO()
                wb2.save(buf2)
                buf2.seek(0)
                response = HttpResponse(
                    buf2.getvalue(),
                    content_type=(
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    ),
                )
                response['Content-Disposition'] = (
                    f'attachment; filename="{stem}_converted.xlsx"'
                )
                return response

            except ImportError:
                return JsonResponse(
                    {'error': 'openpyxl is required for Excel output.'},
                    status=500,
                )


# =============================================================================
# ZONE / SECTOR VIEWS
# =============================================================================

class ProjectZoneListView(LoginRequiredMixin, View):
    """Return all zones as JSON (used by the map page)."""
    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        from apps.projects.models import ProjectZone
        zones = list(project.zones.values('id', 'name', 'color', 'description'))
        return JsonResponse({'zones': zones})


class ProjectZoneCreateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()
        from apps.projects.models import ProjectZone
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Zone name is required.')
            return redirect('project_locations', pk=pk)
        ProjectZone.objects.create(
            project=project,
            name=name,
            color=request.POST.get('color', '#6366f1'),
            description=request.POST.get('description', '').strip(),
        )
        messages.success(request, f'Zone "{name}" created.')
        return redirect('project_locations', pk=pk)


class ProjectZoneEditView(LoginRequiredMixin, View):
    def post(self, request, pk, zid):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()
        from apps.projects.models import ProjectZone
        zone = get_object_or_404(ProjectZone, pk=zid, project=project)
        zone.name        = request.POST.get('name', zone.name).strip()
        zone.color       = request.POST.get('color', zone.color)
        zone.description = request.POST.get('description', zone.description).strip()
        zone.save()
        messages.success(request, f'Zone "{zone.name}" updated.')
        return redirect('project_locations', pk=pk)


class ProjectZoneDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk, zid):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()
        from apps.projects.models import ProjectZone
        zone = get_object_or_404(ProjectZone, pk=zid, project=project)
        zname = zone.name
        # Unlink locations — they stay, just lose their zone
        zone.locations.update(zone=None)
        zone.delete()
        messages.success(request, f'Zone "{zname}" deleted.')
        return redirect('project_locations', pk=pk)


class ProjectLocationBulkZoneView(LoginRequiredMixin, View):
    """
    POST: assign (or un-assign) a list of location IDs to a zone in one shot.
    Body params:
      zone          — zone UUID or empty string (to unassign)
      location_ids  — repeated field, one per selected location
    """
    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        if not _is_project_member(request.user, project):
            return HttpResponseForbidden()

        from apps.projects.models import ProjectZone
        zone_id = request.POST.get('zone', '').strip() or None
        zone    = None
        if zone_id:
            zone = get_object_or_404(ProjectZone, pk=zone_id, project=project)

        loc_ids = request.POST.getlist('location_ids')
        if not loc_ids:
            messages.error(request, 'No locations selected.')
            return redirect('project_locations', pk=pk)

        updated = (
            project.locations
            .filter(pk__in=loc_ids)
            .update(zone=zone)
        )

        if zone:
            messages.success(request, f'{updated} location(s) assigned to "{zone.name}".')
        else:
            messages.success(request, f'{updated} location(s) removed from their zone.')

        return redirect('project_locations', pk=pk)


class IPLocateView(View):
    """
    GET /projects/surveys/ip-locate/
    Returns approximate location for the requester's IP using ipapi.co
    (with ip-api.com fallback). No auth required — public surveys need it.
    """
    http_method_names = ['get']

    def get(self, request):
        # Trust X-Forwarded-For if behind a proxy (same pattern your survey
        # submission endpoint uses to capture ip_address on SurveyResponse).
        xff = request.META.get('HTTP_X_FORWARDED_FOR', '')
        if xff:
            ip = xff.split(',')[0].strip()
        else:
            ip = request.META.get('REMOTE_ADDR', '')

        result = lookup_ip_location(ip)
        if result is None:
            return JsonResponse(
                {'ok': False, 'error': 'IP location lookup failed or unavailable.'},
                status=503,
            )
        return JsonResponse({'ok': True, **result})


class ReverseGeocodeView(View):
    """
    GET /projects/surveys/reverse-geocode/?lat=..&lng=..
    Turns coordinates into a human-readable address (OSM Nominatim).
    Rate-limited server-side.
    """
    http_method_names = ['get']

    def get(self, request):
        try:
            lat = float(request.GET.get('lat'))
            lng = float(request.GET.get('lng'))
        except (TypeError, ValueError):
            return JsonResponse(
                {'ok': False, 'error': 'lat and lng are required floats'},
                status=400,
            )
        if not (-90 <= lat <= 90 and -180 <= lng <= 180):
            return JsonResponse(
                {'ok': False, 'error': 'coordinates out of range'},
                status=400,
            )

        address = reverse_geocode(lat, lng)
        return JsonResponse({'ok': True, 'address': address or ''})